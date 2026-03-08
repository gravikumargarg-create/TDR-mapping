"""Production LVT TDR Delivery view – run from app.py when portal_view == 'production'."""
import io
import tempfile
from datetime import datetime
from pathlib import Path

import streamlit as st

try:
    from lvt_tdr_core import run_lvt_tdr_from_paths, run_tdr_list_only
except ImportError:
    run_lvt_tdr_from_paths = None
    run_tdr_list_only = None


def _normalize_id(val):
    """Normalize BAN/CUSTOMER_ID for comparison (strip, string, handle leading zeros)."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    return s


def _run_capability_validation(excel_bytes):
    """
    Compare QE_BAN_LIST BAN column with Device Details CUSTOMER_ID.
    Returns (missing_bans_list, ban_sheet_name, device_sheet_name, wb) or (None, None, None, None) on error.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return (None, None, None, None)
    try:
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
    except Exception:
        return (None, None, None, None)
    # Find QE_BAN_LIST sheet (or first sheet with BAN column)
    ban_sheet = None
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row < 2:
            continue
        first_row = [ws.cell(1, c).value for c in range(1, min(ws.max_column + 1, 20))]
        if any(first_row and str(x).strip().upper() == "BAN" for x in first_row):
            ban_sheet = name
            break
    if ban_sheet is None:
        return (None, None, None, None)
    # Find Device Details sheet (or last sheet with CUSTOMER_ID)
    device_sheet = None
    for name in reversed(wb.sheetnames):
        ws = wb[name]
        if ws.max_row < 2:
            continue
        first_row = [ws.cell(1, c).value for c in range(1, min(ws.max_column + 1, 20))]
        if any(first_row and "CUSTOMER_ID" in str(x).upper() for x in first_row):
            device_sheet = name
            break
    if device_sheet is None:
        return (None, None, None, None)
    ws_ban = wb[ban_sheet]
    ws_dev = wb[device_sheet]
    # BAN column index (1-based)
    ban_col = None
    for c in range(1, ws_ban.max_column + 1):
        if str(ws_ban.cell(1, c).value or "").strip().upper() == "BAN":
            ban_col = c
            break
    if ban_col is None:
        return (None, None, None, None)
    customer_id_col = None
    for c in range(1, ws_dev.max_column + 1):
        if "CUSTOMER_ID" in str(ws_dev.cell(1, c).value or "").upper():
            customer_id_col = c
            break
    if customer_id_col is None:
        return (None, None, None, None)
    bans_in_list = set()
    for r in range(2, ws_ban.max_row + 1):
        val = _normalize_id(ws_ban.cell(r, ban_col).value)
        if val:
            bans_in_list.add(val)
    customer_ids = set()
    for r in range(2, ws_dev.max_row + 1):
        val = _normalize_id(ws_dev.cell(r, customer_id_col).value)
        if val:
            customer_ids.add(val)
    missing = sorted(bans_in_list - customer_ids, key=lambda x: (len(x), x))
    return (missing, ban_sheet, device_sheet, wb)


def _capability_remove_rows(wb, ban_sheet_name, missing_bans_set):
    """Remove rows from ban sheet where BAN is in missing_bans_set. Returns new workbook bytes."""
    from openpyxl import load_workbook
    ws = wb[ban_sheet_name]
    ban_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or "").strip().upper() == "BAN":
            ban_col = c
            break
    if ban_col is None:
        return None
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        val = _normalize_id(ws.cell(r, ban_col).value)
        if val and val in missing_bans_set:
            rows_to_delete.append(r)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _capability_highlight_rows(wb, ban_sheet_name, missing_bans_set):
    """Highlight rows in ban sheet where BAN is in missing_bans_set. Returns new workbook bytes."""
    from openpyxl.styles import PatternFill
    ws = wb[ban_sheet_name]
    ban_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or "").strip().upper() == "BAN":
            ban_col = c
            break
    if ban_col is None:
        return None
    red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
    for r in range(2, ws.max_row + 1):
        val = _normalize_id(ws.cell(r, ban_col).value)
        if val and val in missing_bans_set:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = red_fill
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def render_production():
    st.markdown(
        """
        <style>
        .stApp { background: #f1f5f9 !important; }
        .block-container { padding: 0.75rem 1rem !important; max-width: 900px !important; }
        section[data-testid="stFileUploader"] { background: #fff !important; border-radius: 8px !important; border-top: 3px solid #0d9488 !important; min-height: 72px !important; padding: 0.4rem 0.6rem !important; }
        section[data-testid="stFileUploader"] [data-testid="stFileUploader"] { min-height: 60px !important; }
        .stButton > button[kind="primary"] { background: #0d9488 !important; color: #fff !important; border-radius: 999px !important; padding: 0.4rem 1rem !important; font-size: 0.9rem !important; }
        div[data-testid="stDownloadButton"] > button { border-radius: 8px !important; border: 1px solid #0d9488 !important; color: #0d9488 !important; }
        [data-testid="stVerticalBlock"] > div { padding-top: 0.25rem !important; padding-bottom: 0.25rem !important; }
        .compact-banner { background: linear-gradient(90deg, #0f766e 0%, #0d9488 50%, #14b8a6 100%); color: #fff; padding: 10px 14px; border-radius: 8px; margin-bottom: 10px; text-align: center; }
        .compact-banner .title { font-size: 1.1rem; font-weight: 700; }
        .compact-banner .sub { font-size: 0.75rem; opacity: 0.95; }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div class="compact-banner">
            <div class="title">Bulk data mapping</div>
            <div class="sub">LVT + data → report & INSERT SQL for BAN Master table.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.sidebar:
        mode = st.selectbox(
            "Choose workflow",
            options=["full", "tdr_only"],
            format_func=lambda x: "Full bulk loading" if x == "full" else "Only TDR customer list analysis",
            key="production_mode",
            help="Pick one below — Full bulk loading: LVT + data → report & INSERT SQL. Only TDR customer list analysis: data files → TDR-wise list (no LVT).",
        )
    # Clear the other mode's result when switching
    if mode == "full" and "tdr_list_result" in st.session_state:
        st.session_state.pop("tdr_list_result", None)
    elif mode == "tdr_only" and "lvt_result" in st.session_state:
        st.session_state.pop("lvt_result", None)

    st.markdown("<div style='margin-bottom: 6px;'></div>", unsafe_allow_html=True)

    if mode == "tdr_only":
        # ----- Only TDR customer list analysis -----
        st.markdown("**Upload data files**")
        _tdr_clear = st.session_state.get("data_clear_tdr", 0)
        data_files = st.file_uploader("Data Excel files (multiple)", type=["xlsx", "xlsm"], accept_multiple_files=True, key=f"data_prod_tdr_{_tdr_clear}", help="TDR data, Rate Plan, etc. No LVT needed.")
        _c1, _c2, _c3 = st.columns([2, 1, 1])
        with _c3:
            if st.button("Clear data files", key="clear_tdr_btn", type="secondary", use_container_width=True, help="Remove all data files and start over"):
                st.session_state["data_clear_tdr"] = _tdr_clear + 1
                st.rerun()
        tdr_only_clicked = st.button("Get TDR Customer List", key="tdr_only_btn", type="primary", use_container_width=True, help="Build TDR-wise customer list from uploaded data files.")

        if tdr_only_clicked and run_tdr_list_only is not None:
            if not data_files or all(f.size == 0 for f in data_files):
                st.warning("Upload at least one **data Excel** file, then click **Get TDR Customer List**.")
            else:
                tmpdir = Path(tempfile.mkdtemp(prefix="tdr_list_streamlit_"))
                try:
                    data_paths = []
                    for i, uf in enumerate(data_files):
                        if uf.size == 0:
                            continue
                        name = uf.name or f"data_{i}.xlsx"
                        p = tmpdir / name
                        p.write_bytes(uf.getvalue())
                        data_paths.append(p)
                    out_path = tmpdir / "TDR_Customer_List.xlsx"
                    log_lines = []
                    def log_fn(msg):
                        log_lines.append(msg)
                    with st.spinner("Building TDR-wise customer list…"):
                        result_path = run_tdr_list_only(data_paths, out_path, log_fn=log_fn)
                    if result_path and result_path.is_file():
                        st.session_state["tdr_list_result"] = {
                            "bytes": result_path.read_bytes(),
                            "name": result_path.name,
                        }
                        st.rerun()
                    else:
                        st.warning("No customer IDs found in the data files.")
                except Exception as e:
                    st.exception(e)
                finally:
                    try:
                        import shutil
                        shutil.rmtree(tmpdir, ignore_errors=True)
                    except Exception:
                        pass
        if tdr_only_clicked and run_tdr_list_only is None:
            st.error("Could not load TDR list module.")

        if "tdr_list_result" in st.session_state:
            r = st.session_state["tdr_list_result"]
            st.success("TDR Customer List ready — download below.")
            st.download_button(
                "⬇ Download TDR Customer List (Excel)",
                data=r["bytes"],
                file_name=r.get("name", "TDR_Customer_List.xlsx"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_tdr_list",
                type="primary",
                use_container_width=True,
            )
        return

    # ----- Full bulk loading (two columns to fit on one screen) -----
    col_lvt, col_data = st.columns(2)
    with col_lvt:
        st.markdown("**1. LVT report**")
        lvt_file = st.file_uploader("LVT Excel", type=["xlsx", "xlsm"], key="lvt_prod", help="Excel with BAN/customer IDs and Pass/Fail status.")
        lvt_sheet = st.text_input("LVT sheet name", value="BAN Wise Result", key="lvt_sheet_prod", help="Sheet in LVT Excel with BAN-wise results.")
    with col_data:
        st.markdown("**2. Data Excel files**")
        _full_clear = st.session_state.get("data_clear_full", 0)
        data_files = st.file_uploader("Data Excel files (multiple)", type=["xlsx", "xlsm"], accept_multiple_files=True, key=f"data_prod_full_{_full_clear}", help="TDR data, Rate Plan, etc. — all non-LVT Excel files.")
        if st.button("Clear data files", key="clear_full_btn", type="secondary", use_container_width=True, help="Remove all data files and upload different ones"):
            st.session_state["data_clear_full"] = _full_clear + 1
            st.rerun()

    with st.expander("Optional – INSERT SQL (OWNER / REQUESTOR / Default TDR)", expanded=False):
        st.caption("Leave blank for empty OWNER/REQUESTOR; Default TDR used for Found-but-no-TDR rows.")
        st.text_input("OWNER (for SQL)", value="", key="owner_prod", help="Value for OWNER column in INSERT SQL.")
        st.text_input("REQUESTOR (for SQL)", value="", key="requestor_prod", help="Value for REQUESTOR column in INSERT SQL.")
        st.text_input("Default TDR", value="", key="default_tdr_prod", help="TDR used when row is Found but has no TDR.")

    run = st.button("Run LVT TDR", type="primary", help="Process LVT + data files → report and INSERT SQL.")

    st.markdown("---")
    st.markdown("**3. Capability validation**")
    st.markdown(
        """
        <div style="padding: 0.6rem 1rem; background: #f0fdfa; border: 1px solid #0d9488; border-radius: 8px; margin-bottom: 0.5rem; font-size: 0.85rem;">
            <strong style="color: #0f766e;">Compare BAN list vs Device Details</strong> — find BANs not in Device Details; then <b>Remove from BAN list</b> or <b>Highlight rows</b> and download.
        </div>
        """,
        unsafe_allow_html=True,
    )
    cap_file = st.file_uploader("BAN list Excel", type=["xlsx", "xlsm"], key="cap_validation_file", help="Excel with QE_BAN_LIST sheet (BAN column) and Device Details sheet (CUSTOMER_ID).")
    cap_run = st.button("Run capability validation", key="cap_validation_run", type="secondary", help="Find BANs in list that are not in Device Details.")

    if cap_run and cap_file and cap_file.size > 0:
        excel_bytes = cap_file.getvalue()
        with st.spinner("Comparing BAN list with Device Details…"):
            missing, ban_sheet, device_sheet, wb = _run_capability_validation(excel_bytes)
        if missing is None:
            st.error("Could not run validation. Ensure the Excel has a sheet with a **BAN** column (e.g. QE_BAN_LIST) and a sheet with **CUSTOMER_ID** (e.g. Device Details).")
        elif not missing:
            st.success("All BANs in the BAN list are present in Device Details. No action needed.")
        else:
            st.session_state["cap_validation_result"] = {
                "missing_bans": missing,
                "ban_sheet": ban_sheet,
                "device_sheet": device_sheet,
                "excel_bytes": excel_bytes,
                "original_name": cap_file.name,
            }
            st.rerun()

    if "cap_validation_result" in st.session_state:
        r = st.session_state["cap_validation_result"]
        missing = r["missing_bans"]
        st.warning(f"**{len(missing)} BAN(s)** in the BAN list are **not** in Device Details sheet.")
        with st.expander("View BANs not in Device Details", expanded=True):
            st.write(", ".join(missing[:50]))
            if len(missing) > 50:
                st.caption(f"… and {len(missing) - 50} more.")
        st.markdown("**Choose an action:**")
        base = Path(r["original_name"]).stem
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Pre-compute both files once and cache (so we don't recompute every rerun)
        _cache_key = (r["original_name"], len(missing))
        if "cap_removed_bytes" not in st.session_state or st.session_state.get("cap_validation_key") != _cache_key:
            from openpyxl import load_workbook
            wb_rem = load_workbook(io.BytesIO(r["excel_bytes"]), data_only=True)
            wb_hi = load_workbook(io.BytesIO(r["excel_bytes"]), data_only=True)
            missing_set = set(missing)
            st.session_state["cap_removed_bytes"] = _capability_remove_rows(wb_rem, r["ban_sheet"], missing_set)
            st.session_state["cap_highlighted_bytes"] = _capability_highlight_rows(wb_hi, r["ban_sheet"], missing_set)
            st.session_state["cap_validation_key"] = _cache_key
        col_remove, col_highlight = st.columns(2)
        with col_remove:
            if st.session_state.get("cap_removed_bytes"):
                st.download_button(
                    "Remove from BAN list and download",
                    data=st.session_state["cap_removed_bytes"],
                    file_name=f"{base}_BANs_removed_{ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="cap_dl_removed",
                    type="primary",
                    use_container_width=True,
                    help="Download Excel with missing BANs removed from the list.",
                )
        with col_highlight:
            if st.session_state.get("cap_highlighted_bytes"):
                st.download_button(
                    "Highlight rows and download",
                    data=st.session_state["cap_highlighted_bytes"],
                    file_name=f"{base}_highlighted_{ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="cap_dl_highlighted",
                    type="secondary",
                    use_container_width=True,
                    help="Download Excel with missing BAN rows highlighted in red.",
                )

    if run and run_lvt_tdr_from_paths is None:
        st.error("Could not load LVT TDR module. Ensure lvt_tdr_core.py is in the app folder.")
    elif run:
        st.session_state.pop("tdr_list_result", None)  # clear TDR-only result when doing full run
        if not lvt_file or lvt_file.size == 0:
            st.warning("Please upload the **LVT Excel** file.")
        elif not data_files or all(f.size == 0 for f in data_files):
            st.warning("Please upload at least one **data Excel** file.")
        else:
            tmpdir = Path(tempfile.mkdtemp(prefix="lvt_tdr_streamlit_"))
            try:
                lvt_path = tmpdir / "lvt_input.xlsx"
                lvt_path.write_bytes(lvt_file.getvalue())
                data_paths = []
                for i, uf in enumerate(data_files):
                    if uf.size == 0:
                        continue
                    name = uf.name or f"data_{i}.xlsx"
                    data_paths.append(tmpdir / name)
                    data_paths[-1].write_bytes(uf.getvalue())
                out_dir = tmpdir / "report"
                out_dir.mkdir(parents=True, exist_ok=True)
                log_lines = []
                def log_fn(msg):
                    log_lines.append(msg)
                owner = st.session_state.get("owner_prod", "") or ""
                requestor = st.session_state.get("requestor_prod", "") or ""
                default_tdr = st.session_state.get("default_tdr_prod", "") or ""
                with st.spinner("Processing…"):
                    report_path, sql_synth_path, sql_prod_path, summary = run_lvt_tdr_from_paths(
                        lvt_path, data_paths, out_dir,
                        lvt_sheet_name=lvt_sheet or "BAN Wise Result",
                        owner=owner.strip() or None,
                        requestor=requestor.strip() or None,
                        default_tdr_id=default_tdr.strip() or None,
                        log_fn=log_fn,
                        log_paths=False,
                    )
                # Format log: never show file paths; always show friendly "download below" messages
                import re
                def _format_log_line(line):
                    s = line.strip()
                    # Core can send short messages (when log_paths=False) or path messages
                    if "Report ready for download" in line:
                        return ("✓ **Report ready** — download using the button below.", "success")
                    if "INSERT SQL ready" in line and "for download" in line:
                        m = re.search(r"\((\d+) statements\)", line)
                        n = m.group(1) if m else ""
                        return (f"✓ **INSERT SQL ready** ({n} statements) — download using the button below." if n else "✓ **INSERT SQL ready** — download using the button below.", "success")
                    # Hide any line that contains a file path (Report or SQL)
                    if "Report:" in line and (".xlsx" in line or "/" in line or "\\" in line or "report" in line.lower() or "tmp" in line):
                        return ("✓ **Report ready** — download using the button below.", "success")
                    if "Wrote " in line and " INSERT " in line and (" to " in line or ".sql" in line):
                        m = re.search(r"Wrote (\d+) INSERT", line)
                        n = m.group(1) if m else "0"
                        return (f"✓ **INSERT SQL ready** ({n} statements) — download using the button below.", "success")
                    if s.startswith("Step "):
                        return (s, "step")
                    if line.startswith("  ") and ".xlsx" in line:
                        return (s, "file")
                    return (line, "text")

                formatted = [_format_log_line(ln) for ln in log_lines]

                with st.expander("**Run log**", expanded=True):
                    st.markdown(
                        """
                        <div style="
                            background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
                            border: 1px solid #cbd5e1;
                            border-radius: 10px;
                            padding: 1.25rem 1.5rem;
                            font-size: 0.9rem;
                            margin-bottom: 0.5rem;
                            box-shadow: 0 1px 2px rgba(0,0,0,0.05);
                        ">
                        """,
                        unsafe_allow_html=True,
                    )
                    for line, kind in formatted:
                        if not line.strip():
                            continue
                        if kind == "step":
                            st.markdown(f"**{line}**")
                        elif kind == "file":
                            st.markdown(f"- {line}")
                        elif kind == "success":
                            st.markdown(f"🟢 {line}")
                        else:
                            st.markdown(line)
                    st.markdown(
                        "<div style='margin-top: 1rem; padding-top: 0.75rem; border-top: 1px solid #e2e8f0; color: #64748b; font-size: 0.85rem;'>Report and SQL are not saved to disk — download them using the buttons below.</div>",
                        unsafe_allow_html=True,
                    )

                if report_path and report_path.is_file():
                    report_bytes = report_path.read_bytes()
                    sql_synth_bytes = sql_synth_path.read_bytes() if sql_synth_path and sql_synth_path.is_file() else None
                    sql_prod_bytes = sql_prod_path.read_bytes() if sql_prod_path and sql_prod_path.is_file() else None
                    st.session_state["lvt_result"] = {
                        "report_bytes": report_bytes,
                        "report_name": report_path.name,
                        "sql_synth_bytes": sql_synth_bytes,
                        "sql_synth_name": sql_synth_path.name if sql_synth_path else None,
                        "sql_prod_bytes": sql_prod_bytes,
                        "sql_prod_name": sql_prod_path.name if sql_prod_path else None,
                        "summary": summary,
                    }
                else:
                    st.error("Report generation failed.")
            except Exception as e:
                st.exception(e)
            finally:
                try:
                    import shutil
                    shutil.rmtree(tmpdir, ignore_errors=True)
                except Exception:
                    pass

    if "lvt_result" in st.session_state:
        r = st.session_state["lvt_result"]
        st.success("Done — download below.")
        c1, c2, c3 = st.columns(3)
        with c1:
            st.download_button(
                "Download report (Excel)",
                data=r["report_bytes"],
                file_name=r["report_name"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_report",
            )
        with c2:
            if r.get("sql_synth_bytes"):
                st.download_button(
                    "INSERT SQL – synthetic data",
                    data=r["sql_synth_bytes"],
                    file_name=r.get("sql_synth_name") or "INSERT_BAN_MASTER_LIST_LVT_SYNTH.sql",
                    mime="text/plain",
                    key="dl_sql_synth",
                )
            else:
                st.info("No INSERT SQL for synthetic (no customers starting with 960).")
        with c3:
            if r.get("sql_prod_bytes"):
                st.download_button(
                    "INSERT SQL – production data",
                    data=r["sql_prod_bytes"],
                    file_name=r.get("sql_prod_name") or "INSERT_BAN_MASTER_LIST_LVT_PRODUCTION.sql",
                    mime="text/plain",
                    key="dl_sql_prod",
                )
            else:
                st.info("No INSERT SQL for production (no customers other than 960*).")

        if r.get("summary"):
            s = r["summary"]
            total = s.get("total", 0)
            passed = s.get("passed", 0)
            failed = s.get("failed", 0)
            not_found = s.get("not_found", 0)
            tdr_p, tdr_f, tdr_part = s.get("tdr_passed", 0), s.get("tdr_failed", 0), s.get("tdr_partial", 0)
            total_tdr = tdr_p + tdr_f + tdr_part

            st.markdown("---")
            st.markdown(
                """
                <div style="
                    background: linear-gradient(135deg, #f0fdfa 0%, #e0f2fe 100%);
                    border: 1px solid #0d9488;
                    border-radius: 10px;
                    padding: 1rem 1.25rem;
                    margin: 0.5rem 0 1rem 0;
                    box-shadow: 0 1px 3px rgba(13, 148, 136, 0.15);
                ">
                    <div style="font-size: 1.1rem; font-weight: 700; color: #0f766e; margin-bottom: 1rem;">High-level summary</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            # BAN wise: only Total, Passed, Failed (no Not found — total = passed + failed)
            ban_html = f"""
            <div style="
                border: 1px solid #0d9488;
                border-radius: 8px;
                padding: 1rem 1.25rem;
                margin-bottom: 1rem;
                background: #fff;
            ">
                <div style="font-weight: 700; color: #0f766e; margin-bottom: 0.75rem; font-size: 1rem;">BAN wise summary</div>
                <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
                    <div style="border: 1px solid #94a3b8; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #f8fafc;">
                        <div style="font-size: 0.8rem; color: #64748b;">Total BAN</div>
                        <div style="font-size: 1.25rem; font-weight: 700;">{total}</div>
                    </div>
                    <div style="border: 1px solid #22c55e; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #dcfce7;">
                        <div style="font-size: 0.8rem; color: #166534;">Passed</div>
                        <div style="font-size: 1.25rem; font-weight: 700; color: #15803d;">{passed}</div>
                    </div>
                    <div style="border: 1px solid #ef4444; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #fee2e2;">
                        <div style="font-size: 0.8rem; color: #991b1b;">Failed</div>
                        <div style="font-size: 1.25rem; font-weight: 700; color: #b91c1c;">{failed}</div>
                    </div>
                </div>
            </div>
            """
            st.markdown(ban_html, unsafe_allow_html=True)

            tdr_html = f"""
            <div style="
                border: 1px solid #0d9488;
                border-radius: 8px;
                padding: 1rem 1.25rem;
                margin-bottom: 0.5rem;
                background: #fff;
            ">
                <div style="font-weight: 700; color: #0f766e; margin-bottom: 0.75rem; font-size: 1rem;">TDR wise summary</div>
                <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
                    <div style="border: 1px solid #94a3b8; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #f8fafc;">
                        <div style="font-size: 0.8rem; color: #64748b;">Total TDR</div>
                        <div style="font-size: 1.25rem; font-weight: 700;">{total_tdr}</div>
                    </div>
                    <div style="border: 1px solid #22c55e; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #dcfce7;">
                        <div style="font-size: 0.8rem; color: #166534;">TDR Passed</div>
                        <div style="font-size: 1.25rem; font-weight: 700; color: #15803d;">{tdr_p}</div>
                    </div>
                    <div style="border: 1px solid #ef4444; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #fee2e2;">
                        <div style="font-size: 0.8rem; color: #991b1b;">TDR Failed</div>
                        <div style="font-size: 1.25rem; font-weight: 700; color: #b91c1c;">{tdr_f}</div>
                    </div>
                    <div style="border: 1px solid #eab308; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #fef9c3;">
                        <div style="font-size: 0.8rem; color: #854d0e;">TDR Partial</div>
                        <div style="font-size: 1.25rem; font-weight: 700; color: #a16207;">{tdr_part}</div>
                    </div>
                </div>
            </div>
            """
            st.markdown(tdr_html, unsafe_allow_html=True)
