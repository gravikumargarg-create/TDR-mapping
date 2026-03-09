"""Synthetic TDR mapping view – run from app.py when portal_view == 'synthetic'."""
import io
import os
import tempfile
import zipfile
from copy import copy
from datetime import datetime
from io import BytesIO
import streamlit as st

import tdr_core
import sharepoint_graph


def _normalize_id(val):
    """Normalize BAN/CUSTOMER_ID for comparison."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _run_tdr_bml_merge(device_details_bytes, bml_excel_bytes):
    """
    Compare CUSTOMER_ID TDR-wise: get BANs from BML Excel TDR sheet, keep only those in Device details.
    Build one Excel with: TDR sheet (copy), Pre-load device details (filtered), BML sheet (filtered).
    Returns (output_bytes, None) or (None, error_message).
    """
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        return (None, "openpyxl required")
    try:
        wb_dev = load_workbook(io.BytesIO(device_details_bytes), data_only=False)
        wb_bml = load_workbook(io.BytesIO(bml_excel_bytes), data_only=False)
    except Exception as e:
        return (None, str(e))
    dev_sheet_name = None
    customer_id_col = None
    for name in wb_dev.sheetnames:
        ws = wb_dev[name]
        if ws.max_row < 2:
            continue
        for c in range(1, ws.max_column + 1):
            if "CUSTOMER_ID" in str(ws.cell(1, c).value or "").upper():
                dev_sheet_name = name
                customer_id_col = c
                break
        if dev_sheet_name:
            break
    if not dev_sheet_name or not customer_id_col:
        return (None, "Device details Excel must have a sheet with CUSTOMER_ID column.")
    ws_dev = wb_dev[dev_sheet_name]
    device_customer_ids = set()
    for r in range(2, ws_dev.max_row + 1):
        val = _normalize_id(ws_dev.cell(r, customer_id_col).value)
        if val:
            device_customer_ids.add(val)
    tdr_sheet_name = None
    tdr_ban_col = None
    tdr_header_row = 1
    for name in wb_bml.sheetnames:
        ws = wb_bml[name]
        if ws.max_row < 2:
            continue
        for row_idx in (2, 1):
            for c in range(1, min(ws.max_column + 1, 30)):
                if str(ws.cell(row_idx, c).value or "").strip().upper() == "BAN":
                    tdr_sheet_name = name
                    tdr_ban_col = c
                    tdr_header_row = row_idx
                    break
            if tdr_sheet_name:
                break
        if tdr_sheet_name:
            break
    if not tdr_sheet_name or not tdr_ban_col:
        return (None, "BML Excel must have a TDR sheet with BAN column.")
    ws_tdr = wb_bml[tdr_sheet_name]
    tdr_bans = set()
    for r in range(tdr_header_row + 1, ws_tdr.max_row + 1):
        val = ws_tdr.cell(r, tdr_ban_col).value
        if val is None:
            continue
        for part in str(val).replace("\r", "\n").split("\n"):
            bid = _normalize_id(part.strip() or part)
            if bid:
                tdr_bans.add(bid)
    valid_bans = tdr_bans & device_customer_ids
    bml_sheet_name = None
    bml_ban_col = None
    for name in wb_bml.sheetnames:
        if name.upper() == "BML":
            bml_sheet_name = name
            ws_bml = wb_bml[name]
            for c in range(1, ws_bml.max_column + 1):
                if str(ws_bml.cell(1, c).value or "").strip().upper() == "BAN":
                    bml_ban_col = c
                    break
            break
    if not bml_sheet_name or not bml_ban_col:
        return (None, "BML Excel must have a sheet named 'BML' with BAN column.")
    preload_name = "Pre-load device details"
    out = Workbook()
    out.remove(out.active)
    ws_tdr_src = wb_bml[tdr_sheet_name]
    ws_tdr_out = out.create_sheet(tdr_sheet_name, 0)
    for r in range(1, ws_tdr_src.max_row + 1):
        for c in range(1, ws_tdr_src.max_column + 1):
            cell_src = ws_tdr_src.cell(r, c)
            cell_out = ws_tdr_out.cell(r, c)
            cell_out.value = cell_src.value
            if cell_src.has_style:
                cell_out.font = copy(cell_src.font)
                cell_out.border = copy(cell_src.border)
                cell_out.fill = copy(cell_src.fill)
                cell_out.number_format = cell_src.number_format
                cell_out.alignment = copy(cell_src.alignment)
    if ws_tdr_src.column_dimensions:
        for k, cd in ws_tdr_src.column_dimensions.items():
            if cd.width:
                ws_tdr_out.column_dimensions[k].width = cd.width
    ws_dev_src = wb_dev[dev_sheet_name]
    ws_preload = out.create_sheet(preload_name, 1)
    for c in range(1, ws_dev_src.max_column + 1):
        cell_src = ws_dev_src.cell(1, c)
        cell_out = ws_preload.cell(1, c)
        cell_out.value = cell_src.value
        if cell_src.has_style:
            cell_out.font = copy(cell_src.font)
            cell_out.border = copy(cell_src.border)
            cell_out.fill = copy(cell_src.fill)
            cell_out.number_format = cell_src.number_format
            cell_out.alignment = copy(cell_src.alignment)
    out_row = 2
    for r in range(2, ws_dev_src.max_row + 1):
        cid = _normalize_id(ws_dev_src.cell(r, customer_id_col).value)
        if cid not in valid_bans:
            continue
        for c in range(1, ws_dev_src.max_column + 1):
            cell_src = ws_dev_src.cell(r, c)
            cell_out = ws_preload.cell(out_row, c)
            cell_out.value = cell_src.value
            if cell_src.has_style:
                cell_out.font = copy(cell_src.font)
                cell_out.border = copy(cell_src.border)
                cell_out.fill = copy(cell_src.fill)
                cell_out.number_format = cell_src.number_format
                cell_out.alignment = copy(cell_src.alignment)
        out_row += 1
    ws_bml_src = wb_bml[bml_sheet_name]
    ws_bml_out = out.create_sheet("BML", 2)
    for c in range(1, ws_bml_src.max_column + 1):
        cell_src = ws_bml_src.cell(1, c)
        cell_out = ws_bml_out.cell(1, c)
        cell_out.value = cell_src.value
        if cell_src.has_style:
            cell_out.font = copy(cell_src.font)
            cell_out.border = copy(cell_src.border)
            cell_out.fill = copy(cell_src.fill)
            cell_out.number_format = cell_src.number_format
            cell_out.alignment = copy(cell_src.alignment)
    out_row = 2
    for r in range(2, ws_bml_src.max_row + 1):
        ban_val = _normalize_id(ws_bml_src.cell(r, bml_ban_col).value)
        if ban_val not in valid_bans:
            continue
        for c in range(1, ws_bml_src.max_column + 1):
            cell_src = ws_bml_src.cell(r, c)
            cell_out = ws_bml_out.cell(out_row, c)
            cell_out.value = cell_src.value
            if cell_src.has_style:
                cell_out.font = copy(cell_src.font)
                cell_out.border = copy(cell_src.border)
                cell_out.fill = copy(cell_src.fill)
                cell_out.number_format = cell_src.number_format
                cell_out.alignment = copy(cell_src.alignment)
        out_row += 1
    buf = io.BytesIO()
    out.save(buf)
    buf.seek(0)
    return (buf.getvalue(), None)


def render_synthetic():
    st.markdown(
        """
        <style>
        .stApp { background: #f1f5f9 !important; }
        .block-container { padding: 1.75rem 1.5rem !important; max-width: 880px !important; }
        section[data-testid="stFileUploader"] { background: #fff !important; border-radius: 8px !important; border-top: 3px solid #0d9488 !important; }
        .stButton > button[kind="primary"] { background: #0d9488 !important; color: #fff !important; border-radius: 999px !important; }
        div[data-testid="stDownloadButton"] > button { border-radius: 8px !important; border: 1px solid #0d9488 !important; color: #0d9488 !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div style="background: linear-gradient(90deg, #0f766e 0%, #0d9488 50%, #14b8a6 100%); color: #fff; padding: 18px 20px; border-radius: 10px; margin-bottom: 12px; text-align: center;">
            <div style="font-size: 1.25rem; font-weight: 700;">TDR wise mapping</div>
            <div style="font-size: 0.75rem; opacity: 0.95;">TDR-wise mapping for synthetic data. Inputs needed: TDR data sheets, device details, and LVT report.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    TDR_SHAREPOINT_URL = (
        "https://amdocs.sharepoint.com/:f:/r/sites/USCCTesting_Offshore/Shared%20Documents/"
        "Release%20%26%20OffCycle/UTMO%20-%20Migration%20(Data%20creation)/Data%20Creation/R2%20Data"
        "?csf=1&web=1&e=EZDfgA"
    )
    tdr_bytes = None
    tdr_name = None
    tdr_sheet = None
    tdr_source = st.radio("TDR file from", options=["Local file", "SharePoint"], horizontal=True, key="tdr_source")
    use_sharepoint_direct = tdr_source == "SharePoint" and sharepoint_graph.has_sharepoint_credentials()
    tdr_bytes = None
    tdr_name = None
    lvt_file = None
    lvt_sheet = None
    device_file = None
    bml_file = None

    if tdr_source == "SharePoint" and not use_sharepoint_direct:
        st.markdown(f'<a href="{TDR_SHAREPOINT_URL}" target="_blank" style="font-size: 0.85rem; color: #0d9488;">📂 Open TDR folder on SharePoint</a>', unsafe_allow_html=True)
        st.caption("Download the file from SharePoint, then use the upload options below.")

    if use_sharepoint_direct:
        token = sharepoint_graph.get_token()
        if not token:
            st.warning("Could not get SharePoint token. Check AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET in secrets.")
        else:
            files = sharepoint_graph.list_tdr_excel_files(token)
            if not files:
                st.info("No Excel files found in the TDR folder.")
            else:
                selected_name = st.selectbox("1. TDR Data (SharePoint)", options=[f["name"] for f in files], key="sp_tdr_file")
                selected = next((f for f in files if f["name"] == selected_name), None)
                if selected:
                    cache_key = f"sp_tdr_{selected['id']}"
                    if cache_key not in st.session_state:
                        with st.spinner("Loading file…"):
                            content = sharepoint_graph.download_file_content(token, selected["drive_id"], selected["id"])
                        if content:
                            st.session_state[cache_key] = content
                    content = st.session_state.get(cache_key) if cache_key in st.session_state else None
                    if content:
                        tdr_bytes = content
                        tdr_name = selected["name"]
                        try:
                            wb_tdr = tdr_core.load_workbook(BytesIO(content), read_only=True, data_only=True)
                            roles = tdr_core.detect_excel_roles(wb_tdr)
                            all_sheets_sp = wb_tdr.sheetnames
                            wb_tdr.close()
                            tdr_sheet_list = roles.get("tdr_sheets") or all_sheets_sp
                            if tdr_sheet_list:
                                st.caption(f"TDR Data → {tdr_name} ({len(tdr_sheet_list)} sheet(s))")
                        except Exception as e:
                            st.warning(str(e))
        st.markdown("**2. LVT file** | **3. Device details** | **4. BML file**")
        _s1, _s2, _s3 = st.columns(3)
        with _s1:
            lvt_file = st.file_uploader("LVT file", type=["xlsx", "xlsm"], key="lvt_upload_sp", help="Excel with BAN Wise Result sheet.")
        with _s2:
            device_file = st.file_uploader("Device details", type=["xlsx", "xlsm"], key="device_upload_sp", help="Excel with CUSTOMER_ID column.")
        with _s3:
            bml_file = st.file_uploader("BML file", type=["xlsx", "xlsm"], key="bml_upload_sp", help="BML Excel (TDR + BML sheets).")
        if lvt_file and lvt_file.size > 0:
            try:
                wb_lvt = tdr_core.load_workbook(BytesIO(lvt_file.getvalue()), read_only=True, data_only=True)
                roles_lvt = tdr_core.detect_excel_roles(wb_lvt)
                wb_lvt.close()
                lvt_sheet_list = roles_lvt.get("lvt_sheets") or []
                lvt_sheet = tdr_core.LVT_SHEET_NAME if tdr_core.LVT_SHEET_NAME in (roles_lvt.get("lvt_sheets") or []) else (lvt_sheet_list[0] if lvt_sheet_list else tdr_core.LVT_SHEET_NAME)
            except Exception:
                lvt_sheet = tdr_core.LVT_SHEET_NAME
    else:
        tdr_sheet_list = []
        st.markdown("**1. Data details input file** | **2. LVT file**")
        _r1a, _r1b = st.columns(2)
        with _r1a:
            data_details_file = st.file_uploader("Data details input file", type=["xlsx", "xlsm"], key="data_details_upload", help="TDR data Excel (all sheets will be processed).")
        with _r1b:
            lvt_file = st.file_uploader("LVT file", type=["xlsx", "xlsm"], key="lvt_upload", help="Excel with BAN Wise Result sheet.")
        st.markdown("**3. Device details** | **4. BML file**")
        _r2a, _r2b = st.columns(2)
        with _r2a:
            device_file = st.file_uploader("Device details", type=["xlsx", "xlsm"], key="device_upload", help="Excel with CUSTOMER_ID column. Added to ZIP when provided.")
        with _r2b:
            bml_file = st.file_uploader("BML file", type=["xlsx", "xlsm"], key="bml_upload", help="BML Excel. Added to ZIP when provided.")
        if data_details_file and data_details_file.size > 0:
            tdr_bytes = data_details_file.getvalue()
            tdr_name = data_details_file.name
            try:
                wb_tdr = tdr_core.load_workbook(BytesIO(tdr_bytes), read_only=True, data_only=True)
                roles = tdr_core.detect_excel_roles(wb_tdr)
                all_sheets = wb_tdr.sheetnames
                wb_tdr.close()
                tdr_sheet_list = roles.get("tdr_sheets") or all_sheets
                st.caption(f"Data details → {tdr_name} ({len(tdr_sheet_list)} sheet(s) will be processed)")
            except Exception:
                tdr_sheet_list = []
        if lvt_file and lvt_file.size > 0:
            try:
                wb_lvt = tdr_core.load_workbook(BytesIO(lvt_file.getvalue()), read_only=True, data_only=True)
                roles_lvt = tdr_core.detect_excel_roles(wb_lvt)
                wb_lvt.close()
                lvt_sheet_list = roles_lvt.get("lvt_sheets") or []
                lvt_sheet = tdr_core.LVT_SHEET_NAME if tdr_core.LVT_SHEET_NAME in (roles_lvt.get("lvt_sheets") or []) else (lvt_sheet_list[0] if lvt_sheet_list else tdr_core.LVT_SHEET_NAME)
            except Exception:
                lvt_sheet = tdr_core.LVT_SHEET_NAME

    st.markdown("<div style='height: 12px;'></div>", unsafe_allow_html=True)
    _run1, _run2, _run3 = st.columns([1, 2, 1])
    with _run2:
        has_tdr = bool(tdr_bytes)
        has_lvt = lvt_file and lvt_file.size > 0
        run = st.button("Run TDR", type="primary", use_container_width=True, disabled=not (has_tdr and has_lvt))

    st.markdown("---")
    st.markdown(
        """
        <span style="font-weight: 700;">TDR + BML merge</span>
        <span title="Upload Device details (CUSTOMER_ID) and BML Excel (TDR sheet + BML sheet). We compare CUSTOMER_ID TDR-wise and build one Excel: TDR sheet, Pre-load device details (filtered), BML sheet (filtered). Same format as reference." style="cursor: help; margin-left: 4px; opacity: 0.8;">ⓘ</span>
        """,
        unsafe_allow_html=True,
    )
    _c4a, _c4b = st.columns(2)
    with _c4a:
        tdr_bml_device_file = st.file_uploader("Device details Excel", type=["xlsx", "xlsm"], key="tdr_bml_device", help="Excel with a sheet containing CUSTOMER_ID column (e.g. Pre-load device details).")
    with _c4b:
        tdr_bml_bml_file = st.file_uploader("BML Excel", type=["xlsx", "xlsm"], key="tdr_bml_bml", help="Excel with TDR sheet (BAN column) and BML sheet (BAN column). Same format as TDR-200581.xlsx.")
    _c4_1, _c4_2, _c4_3 = st.columns([1, 2, 1])
    with _c4_2:
        tdr_bml_run = st.button("Generate TDR Excel", key="tdr_bml_run", type="secondary", use_container_width=True, help="Compare CUSTOMER_ID TDR-wise; output one Excel with TDR, Pre-load device details, BML sheets.")
    if tdr_bml_run and tdr_bml_device_file and tdr_bml_bml_file and tdr_bml_device_file.size > 0 and tdr_bml_bml_file.size > 0:
        with st.spinner("Comparing CUSTOMER_ID TDR-wise and building Excel…"):
            out_bytes, err = _run_tdr_bml_merge(tdr_bml_device_file.getvalue(), tdr_bml_bml_file.getvalue())
        if err:
            st.error(err)
        else:
            st.session_state["tdr_bml_result"] = {"bytes": out_bytes, "name": f"TDR_BML_merged_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
            st.rerun()
    if "tdr_bml_result" in st.session_state:
        r = st.session_state["tdr_bml_result"]
        st.success("TDR Excel ready — download below.")
        st.download_button(
            "Download TDR Excel (TDR + Pre-load device details + BML)",
            data=r["bytes"],
            file_name=r["name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="tdr_bml_dl",
            type="primary",
            use_container_width=True,
        )

    if run and not tdr_bytes:
        st.warning("Please provide **TDR Data** (upload file(s) or pick from SharePoint).")
    elif run and (not lvt_file or lvt_file.size == 0):
        st.warning("Please upload file(s) that include **LVT Report** (sheet named BAN Wise Result).")
    elif run and tdr_bytes and lvt_file and lvt_file.size > 0:
        tmpdir = tempfile.mkdtemp(prefix="tdr_streamlit_")
        try:
            wb_tdr = tdr_core.load_workbook(BytesIO(tdr_bytes), read_only=True, data_only=True)
            roles = tdr_core.detect_excel_roles(wb_tdr)
            all_sheet_names = wb_tdr.sheetnames
            wb_tdr.close()
            sheet_list = roles.get("tdr_sheets") or all_sheet_names
            if not sheet_list:
                st.error("No TDR sheets found in the TDR Data Excel.")
            else:
                os.environ["TDR_WEB_REPORT_FOLDER"] = tmpdir
                tdr_path = os.path.join(tmpdir, "tdr_input.xlsx")
                with open(tdr_path, "wb") as f:
                    f.write(tdr_bytes)
                lvt_path = os.path.join(tmpdir, "lvt_input.xlsx")
                with open(lvt_path, "wb") as f:
                    f.write(lvt_file.getvalue())
                device_details_path = None
                if device_file and device_file.size > 0:
                    device_details_path = os.path.join(tmpdir, "device_details.xlsx")
                    with open(device_details_path, "wb") as f:
                        f.write(device_file.getvalue())
                sheet_to_use = (lvt_sheet or tdr_core.LVT_SHEET_NAME).strip() or tdr_core.LVT_SHEET_NAME
                out_path = os.path.join(tmpdir, "TDR_BAN_Report.xlsx")
                with st.spinner("Processing…"):
                    result_path, summary = tdr_core.run_extraction_and_report(
                        [(tdr_path, sheet_list)], output_excel=out_path,
                        lvt_report_path=lvt_path, lvt_sheet_name=sheet_to_use if lvt_path else None,
                        device_details_path=device_details_path,
                    )
                if result_path and os.path.isfile(result_path):
                    with open(result_path, "rb") as f:
                        report_bytes = f.read()
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    report_filename = f"TDR_BAN_Report_{ts}.xlsx"
                    zip_bytes = None
                    zip_filename = f"TDR_wise_report_{datetime.now().strftime('%Y%m%d')}.zip"
                    per_tdr_folder = (summary or {}).get("per_tdr_folder")
                    if per_tdr_folder and os.path.isdir(per_tdr_folder):
                        files = [n for n in os.listdir(per_tdr_folder) if n.endswith((".xlsx", ".xlsm"))]
                        has_device = device_file and device_file.size > 0 and device_details_path and os.path.isfile(device_details_path)
                        has_bml = bml_file and bml_file.size > 0
                        if files or has_device or has_bml:
                            buf = BytesIO()
                            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
                                for n in files:
                                    z.write(os.path.join(per_tdr_folder, n), n)
                                if has_device:
                                    z.write(device_details_path, "Pre-load device details.xlsx")
                                if has_bml:
                                    bml_path = os.path.join(tmpdir, "bml_input.xlsx")
                                    with open(bml_path, "wb") as f:
                                        f.write(bml_file.getvalue())
                                    z.write(bml_path, "BML.xlsx")
                            buf.seek(0)
                            zip_bytes = buf.getvalue()
                    st.session_state["tdr_result"] = {
                        "report_bytes": report_bytes, "report_filename": report_filename,
                        "zip_bytes": zip_bytes, "zip_filename": zip_filename, "summary": summary, "lvt_used": True,
                    }
                else:
                    st.error("No TDR data found or report generation failed.")
        finally:
            if "TDR_WEB_REPORT_FOLDER" in os.environ:
                del os.environ["TDR_WEB_REPORT_FOLDER"]
            try:
                import shutil
                shutil.rmtree(tmpdir, ignore_errors=True)
            except Exception:
                pass

    if "tdr_result" in st.session_state:
        r = st.session_state["tdr_result"]
        st.success("Done — download below.")
        if r.get("zip_bytes"):
            c1, c2 = st.columns(2)
            with c1:
                st.download_button("Download main report", data=r["report_bytes"], file_name=r["report_filename"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_main")
            with c2:
                st.download_button("Download per-TDR (ZIP)", data=r["zip_bytes"], file_name=r["zip_filename"], mime="application/zip", key="dl_zip")
        else:
            st.download_button("Download main report", data=r["report_bytes"], file_name=r["report_filename"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_main")
        if r.get("summary"):
            s = r["summary"]
            total = s.get("total", 0)
            passed, failed, not_found = s.get("passed", 0), s.get("failed", 0), s.get("not_found", 0)
            tdr_p, tdr_f, tdr_part = s.get("tdr_passed", 0), s.get("tdr_failed", 0), s.get("tdr_partial", 0)
            total_tdr = tdr_p + tdr_f + tdr_part

            st.markdown("---")
            st.markdown(
                """
                <div style="
                    background: linear-gradient(135deg, #f0fdfa 0%, #e0f2fe 100%);
                    border: 1px solid #0d9488;
                    border-radius: 10px;
                    padding: 1rem 1.25rem;
                    margin: 0.5rem 0 1rem 0;
                    box-shadow: 0 1px 3px rgba(13, 148, 136, 0.15);
                ">
                    <div style="font-size: 1.1rem; font-weight: 700; color: #0f766e; margin-bottom: 1rem;">High-level summary</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            # BAN wise summary — only Total, Passed, Failed (no Not found)
            ban_html = f"""
            <div style="
                border: 1px solid #0d9488;
                border-radius: 8px;
                padding: 1rem 1.25rem;
                margin-bottom: 1rem;
                background: #fff;
            ">
                <div style="font-weight: 700; color: #0f766e; margin-bottom: 0.75rem; font-size: 1rem;">BAN wise summary</div>
                <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
                    <div style="border: 1px solid #94a3b8; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #f8fafc;">
                        <div style="font-size: 0.8rem; color: #64748b;">Total BAN</div>
                        <div style="font-size: 1.25rem; font-weight: 700;">{total}</div>
                    </div>
                    <div style="border: 1px solid #22c55e; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #dcfce7;">
                        <div style="font-size: 0.8rem; color: #166534;">Passed</div>
                        <div style="font-size: 1.25rem; font-weight: 700; color: #15803d;">{passed}</div>
                    </div>
                    <div style="border: 1px solid #ef4444; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #fee2e2;">
                        <div style="font-size: 0.8rem; color: #991b1b;">Failed</div>
                        <div style="font-size: 1.25rem; font-weight: 700; color: #b91c1c;">{failed}</div>
                    </div>
                </div>
            </div>
            """
            st.markdown(ban_html, unsafe_allow_html=True)

            # TDR wise summary — bordered section, total first then breakdown with colors
            tdr_html = f"""
            <div style="
                border: 1px solid #0d9488;
                border-radius: 8px;
                padding: 1rem 1.25rem;
                margin-bottom: 0.5rem;
                background: #fff;
            ">
                <div style="font-weight: 700; color: #0f766e; margin-bottom: 0.75rem; font-size: 1rem;">TDR wise summary</div>
                <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
                    <div style="border: 1px solid #94a3b8; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #f8fafc;">
                        <div style="font-size: 0.8rem; color: #64748b;">Total TDR</div>
                        <div style="font-size: 1.25rem; font-weight: 700;">{total_tdr}</div>
                    </div>
                    <div style="border: 1px solid #22c55e; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #dcfce7;">
                        <div style="font-size: 0.8rem; color: #166534;">TDR Passed</div>
                        <div style="font-size: 1.25rem; font-weight: 700; color: #15803d;">{tdr_p}</div>
                    </div>
                    <div style="border: 1px solid #ef4444; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #fee2e2;">
                        <div style="font-size: 0.8rem; color: #991b1b;">TDR Failed</div>
                        <div style="font-size: 1.25rem; font-weight: 700; color: #b91c1c;">{tdr_f}</div>
                    </div>
                    <div style="border: 1px solid #eab308; border-radius: 6px; padding: 0.5rem 1rem; min-width: 90px; background: #fef9c3;">
                        <div style="font-size: 0.8rem; color: #854d0e;">TDR Partial</div>
                        <div style="font-size: 1.25rem; font-weight: 700; color: #a16207;">{tdr_part}</div>
                    </div>
                </div>
            </div>
            """
            st.markdown(tdr_html, unsafe_allow_html=True)
