"""
TDR Data Excel script - Sheet/file selection and TDR → BAN extraction.
- Asks user which sheet to use from TDR Data.xlsx (default: same folder as script).
- Supports multiple Excel files and/or multiple sheets if needed.
- Headings are not static; identifies sections by TDR number (TDR-######,
  TDR ######, TDR_###### or TDR######) and collects 9-digit BAN IDs per section.
"""

import os
import re
import shutil
import sys
from copy import copy
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    # Do not sys.exit(1) when running inside Streamlit (causes "Error running app" with no message)
    if "streamlit" in sys.modules:
        raise ImportError("openpyxl is required. Add 'openpyxl>=3.1.0' to requirements.txt and redeploy.") from e
    print("openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# Pattern: TDR number – hyphen (TDR-204035), space (TDR 203410), underscore (TDR_203410), or none
TDR_PATTERN = re.compile(r"TDR[\s\-_]*(\d{5,6})", re.IGNORECASE)
# Pattern: 9-digit BAN ID (standalone number in string)
BAN_PATTERN = re.compile(r"\b(\d{9})\b")


# Script folder; reports in report subfolder; old reports moved to report/archive
# When TDR_WEB_REPORT_FOLDER env is set (e.g. by web/Cloud Function), use that for output
BASE_FOLDER = os.path.dirname(os.path.abspath(__file__))
REPORT_FOLDER = os.environ.get("TDR_WEB_REPORT_FOLDER") or os.path.join(BASE_FOLDER, "report")
ARCHIVE_FOLDER = os.path.join(REPORT_FOLDER, "archive")
# Console styling (ANSI on Windows 10+)
if sys.platform == "win32":
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)
    except Exception:
        pass
C = type("C", (), {"RESET": "\033[0m", "BOLD": "\033[1m", "GREEN": "\033[92m", "RED": "\033[91m", "YELLOW": "\033[93m", "CYAN": "\033[96m", "DIM": "\033[2m"})()
BOX = {"TL": "╔", "TR": "╗", "BL": "╚", "BR": "╝", "H": "═", "V": "║"}
W = 56
DEFAULT_TDR_EXCEL = os.path.join(BASE_FOLDER, "TDR Data.xlsx")
DEFAULT_LVT_REPORT = os.path.join(BASE_FOLDER, "LVT_RUN_25Feb_Report.xlsx")
LVT_SHEET_NAME = "BAN Wise Result"
# Failures sheet: match any variation of "BAN Wise Failures" (spaces, underscores, case)
LVT_FAILURES_SHEET_NORMALIZED = "banwisefailures"

# Comments for TDR Info column F: check_id (str) -> comment text. Add more as needed.
CHECK_ID_COMMENTS = {
    "3106": "BANs can be used after RS-122932 fix.",
}


def _normalize_sheet_name(name):
    """Return name with spaces/underscores removed and lowercased for matching."""
    if not name:
        return ""
    return "".join(c for c in str(name).lower().strip() if c not in " \t_")


def _find_failures_sheet_in_workbook(wb):
    """Return the first sheet name in wb that normalizes to BAN Wise Failures, or None."""
    target = _normalize_sheet_name(LVT_FAILURES_SHEET_NORMALIZED)
    for sheet_name in wb.sheetnames:
        if _normalize_sheet_name(sheet_name) == target:
            return sheet_name
    return None


def detect_excel_roles(wb):
    """
    Detect which roles this workbook can serve (TDR Data, LVT Report, Device Details).
    Returns dict: tdr_sheets=[], lvt_sheets=[], device_sheets=[] (sheet names).
    """
    result = {"tdr_sheets": [], "lvt_sheets": [], "device_sheets": []}
    lvt_target = _normalize_sheet_name(LVT_SHEET_NAME)  # "banwiseresult"
    for name in wb.sheetnames:
        try:
            ws = wb[name]
        except Exception:
            continue
        try:
            if extract_tdr_ban_mapping(ws):
                result["tdr_sheets"].append(name)
        except Exception:
            pass
        if _normalize_sheet_name(name) == lvt_target:
            result["lvt_sheets"].append(name)
        try:
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row and _find_column_index_in_row(list(first_row), DEVICE_DETAILS_CUSTOMER_ID_HEADERS):
                result["device_sheets"].append(name)
        except Exception:
            pass
    return result
# Folder to scan for LVT Excel files (user picks file and optionally sheet)
TDR_DELIVERY_FOLDER = os.path.join(BASE_FOLDER, "TDR deliver")


def archive_old_reports(report_folder, max_age_days=1):
    """
    Move report files and TDR folders in report_folder that are older than max_age_days
    into report_folder/archive. Creates archive folder if needed.
    """
    if not os.path.isdir(report_folder):
        return
    archive_dir = os.path.join(report_folder, "archive")
    os.makedirs(archive_dir, exist_ok=True)
    cutoff = datetime.now() - timedelta(days=max_age_days)
    cutoff_ts = cutoff.timestamp()
    for name in os.listdir(report_folder):
        if name == "archive":
            continue
        path = os.path.join(report_folder, name)
        try:
            mtime = os.path.getmtime(path)
        except OSError:
            continue
        if mtime >= cutoff_ts:
            continue
        dest = os.path.join(archive_dir, name)
        try:
            if os.path.exists(dest):
                base, ext = os.path.splitext(name) if os.path.isfile(path) else (name, "")
                if ext:
                    n = 1
                    while os.path.exists(dest):
                        dest = os.path.join(archive_dir, f"{base}_{n}{ext}")
                        n += 1
                else:
                    n = 1
                    while os.path.exists(dest):
                        dest = os.path.join(archive_dir, f"{base}_{n}")
                        n += 1
            shutil.move(path, dest)
        except (OSError, shutil.Error):
            pass


def get_excel_path():
    """Ask user for TDR Data Excel path; use default if empty."""
    print(f"\nTDR Data Excel (default: {DEFAULT_TDR_EXCEL})")
    path = input("  Path [Enter for default]: ").strip()
    if not path:
        path = DEFAULT_TDR_EXCEL
    if not os.path.isfile(path):
        print(f"  Error: File not found: {path}")
        return None
    return path


def _excel_extensions():
    return (".xlsx", ".xlsm")


def _list_lvt_excel_files(folder):
    """Return list of (full_path, filename) for files in folder whose name starts with 'LVT' (case-insensitive)."""
    if not os.path.isdir(folder):
        return []
    result = []
    for name in os.listdir(folder):
        if not name.upper().startswith("LVT"):
            continue
        if name.lower().endswith(_excel_extensions()):
            result.append((os.path.join(folder, name), name))
    return sorted(result, key=lambda x: x[1])


def ask_single_sheet_choice(sheet_names, label="workbook"):
    """Ask user to pick one sheet by number or name. Returns sheet name or None."""
    print(f"\nSheets in {label}:")
    for i, name in enumerate(sheet_names, 1):
        print(f"  {i}. {name}")
    print("  Enter sheet number or sheet name to use for BAN-wise list.")
    choice = input("  Your choice: ").strip()
    if not choice:
        return None
    if choice.isdigit():
        idx = int(choice)
        if 1 <= idx <= len(sheet_names):
            return sheet_names[idx - 1]
        print(f"  Invalid number: {idx}")
        return None
    if choice in sheet_names:
        return choice
    print(f"  Sheet not found: {choice!r}")
    return None


def get_lvt_report_file_and_sheet():
    """
    Scan TDR_DELIVERY_FOLDER for Excel files whose name starts with 'LVT'.
    Let user select one file; then if sheet 'BAN Wise Result' exists use it,
    else list all sheets and ask which sheet to use for BAN-wise list.
    Returns (file_path, sheet_name) or None.
    """
    print(f"\nLVT Report – selecting file from delivery path: {TDR_DELIVERY_FOLDER}")
    lvt_files = _list_lvt_excel_files(TDR_DELIVERY_FOLDER)
    if not lvt_files:
        print("  No Excel file starting with 'LVT' found in this folder.")
        print("  Status column will show 'Not found' for all BANs.")
        return None

    print("  LVT Excel files found:")
    for i, (path, name) in enumerate(lvt_files, 1):
        print(f"  {i}. {name}")
    print("  Enter file number or filename to use for BAN-wise list.")
    choice = input("  Your choice: ").strip()
    if not choice:
        print("  No selection. Skipping LVT report.")
        return None

    selected_path = None
    if choice.isdigit():
        idx = int(choice)
        if 1 <= idx <= len(lvt_files):
            selected_path = lvt_files[idx - 1][0]
        else:
            print(f"  Invalid number: {idx}")
            return None
    else:
        for path, name in lvt_files:
            if name == choice or name.lower() == choice.lower():
                selected_path = path
                break
        if not selected_path:
            print(f"  File not found: {choice!r}")
            return None

    # Open workbook and decide sheet
    try:
        wb = load_workbook(selected_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as e:
        print(f"  Error opening file: {e}")
        return None

    if LVT_SHEET_NAME in sheet_names:
        print(f"  Using sheet: {LVT_SHEET_NAME}")
        return (selected_path, LVT_SHEET_NAME)

    print(f"  Expected sheet {LVT_SHEET_NAME!r} not found in this workbook.")
    chosen = ask_single_sheet_choice(sheet_names, os.path.basename(selected_path))
    if not chosen:
        return None
    return (selected_path, chosen)


def get_sheet_names(excel_path):
    """Return list of sheet names in the workbook."""
    wb = load_workbook(excel_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def ask_sheet_choice(sheet_names, excel_label="TDR Data"):
    """Ask user to pick one or more sheets; return list of selected sheet names."""
    print(f"\nSheets in {excel_label}:")
    for i, name in enumerate(sheet_names, 1):
        print(f"  {i}. {name}")
    print("  Enter sheet number(s) separated by comma (e.g. 4 or 1,3,5), or sheet name(s).")
    choice = input("  Your choice: ").strip()
    if not choice:
        return None
    selected = []
    # Allow numbers
    for part in (c.strip() for c in choice.split(",")):
        if part.isdigit():
            idx = int(part)
            if 1 <= idx <= len(sheet_names):
                selected.append(sheet_names[idx - 1])
            else:
                print(f"  Invalid number: {idx}")
        else:
            # Treat as sheet name
            if part in sheet_names:
                selected.append(part)
            else:
                print(f"  Sheet not found: {part!r}")
    return selected if selected else None


def ask_more_sources():
    """Ask if user needs to add more Excel files or more sheets."""
    print("\nDo you need to use another Excel file or more sheets from the same file? (y/n)")
    return input("  Your choice [n]: ").strip().lower() in ("y", "yes")


def _extract_tdr_from_cell(value):
    """If cell value contains TDR-######, return e.g. 'TDR-204035'; else None."""
    if value is None:
        return None
    s = str(value).strip()
    m = TDR_PATTERN.search(s)
    if m:
        return f"TDR-{m.group(1)}"
    return None


def _extract_bans_from_cell(value):
    """Extract all 9-digit BAN IDs from a cell (number or string with newlines). Returns set of strings."""
    bans = set()
    if value is None:
        return bans
    if isinstance(value, (int, float)):
        v = int(value)
        if 100_000_000 <= v <= 999_999_999:
            bans.add(str(v))
        return bans
    s = str(value).strip()
    for m in BAN_PATTERN.finditer(s):
        bans.add(m.group(1))
    return bans


def extract_tdr_ban_mapping(ws):
    """
    Scan sheet: find rows where any cell contains TDR-###### (section header).
    For each section (from one TDR row to just before the next), collect all
    9-digit BAN IDs from any cell. Return dict: TDR_number -> list of BAN IDs (unique).
    """
    # Find all TDR header rows: (row_1_index, tdr_id)
    tdr_rows = []  # list of (row_idx, tdr_id), 1-based
    for row in ws.iter_rows():
        row_idx = None
        for cell in row:
            if hasattr(cell, "row"):
                row_idx = cell.row
                break
        if not row_idx:
            continue
        for cell in row:
            tdr_id = _extract_tdr_from_cell(cell.value)
            if tdr_id:
                tdr_rows.append((row_idx, tdr_id))
                break

    if not tdr_rows:
        return {}

    # For each consecutive pair of TDR rows, collect BANs in that block
    result = {}
    for i, (start_row, tdr_id) in enumerate(tdr_rows):
        end_row = tdr_rows[i + 1][0] if i + 1 < len(tdr_rows) else (ws.max_row + 1)
        bans = set()
        for row in ws.iter_rows(min_row=start_row, max_row=end_row - 1):
            for cell in row:
                bans |= _extract_bans_from_cell(cell.value)
        result[tdr_id] = sorted(bans)  # sorted for consistent output

    return result


def get_tdr_section_ranges(ws):
    """
    Find all TDR section header rows and return (tdr_id, start_row, end_row) for each section.
    Section includes rows from start_row through end_row - 1 (end_row is exclusive).
    """
    tdr_rows = []
    for row in ws.iter_rows():
        row_idx = next((c.row for c in row if hasattr(c, "row")), None)
        if not row_idx:
            continue
        for cell in row:
            tdr_id = _extract_tdr_from_cell(cell.value)
            if tdr_id:
                tdr_rows.append((row_idx, tdr_id))
                break
    if not tdr_rows:
        return []
    result = []
    for i, (start_row, tdr_id) in enumerate(tdr_rows):
        end_row_exclusive = tdr_rows[i + 1][0] if i + 1 < len(tdr_rows) else (ws.max_row + 1)
        result.append((tdr_id, start_row, end_row_exclusive))
    return result


def _copy_cell_style(src_cell, dest_cell):
    """Copy font, fill, border, alignment, number_format from src_cell to dest_cell."""
    try:
        if getattr(src_cell, "font", None):
            dest_cell.font = copy(src_cell.font)
        if getattr(src_cell, "fill", None):
            dest_cell.fill = copy(src_cell.fill)
        if getattr(src_cell, "border", None):
            dest_cell.border = copy(src_cell.border)
        if getattr(src_cell, "alignment", None):
            dest_cell.alignment = copy(src_cell.alignment)
        if getattr(src_cell, "number_format", None):
            dest_cell.number_format = src_cell.number_format
        if getattr(src_cell, "protection", None):
            dest_cell.protection = copy(src_cell.protection)
    except Exception:
        pass


def _is_cell_non_empty(value):
    """Return True if value is considered non-empty (for used-range detection)."""
    if value is None:
        return False
    s = str(value).strip()
    return len(s) > 0


def _get_used_column_range(src_ws, start_row, end_row_exclusive, merged_ranges):
    """
    For the row range [start_row, end_row_exclusive), find the min and max column
    that contain at least one non-empty cell (merge-aware). Returns (min_col, max_col), 1-based.
    """
    min_col, max_col = None, None
    for r in range(start_row, end_row_exclusive):
        for c in range(1, src_ws.max_column + 1):
            val = _get_cell_value_respecting_merge(src_ws, r, c, merged_ranges)
            if _is_cell_non_empty(val):
                if min_col is None:
                    min_col = max_col = c
                else:
                    min_col = min(min_col, c)
                    max_col = max(max_col, c)
    if min_col is None or max_col is None:
        return (1, max(1, src_ws.max_column))
    return (min_col, max_col)


def _copy_sheet_range_to_workbook(src_ws, start_row, end_row_exclusive, sheet_title="Data"):
    """
    Copy all rows from src_ws (start_row through end_row_exclusive - 1) and all used columns
    (columns that have at least one non-empty cell in this range) into a new Workbook.
    Respects merged cells (uses top-left value for merged area).
    Copies values, column widths, row heights, and cell formatting.
    """
    wb = Workbook()
    dest_ws = wb.active
    dest_ws.title = sheet_title[:31] if sheet_title else "Data"
    merged_ranges = _get_merged_ranges(src_ws)
    min_col, max_col = _get_used_column_range(src_ws, start_row, end_row_exclusive, merged_ranges)
    for r in range(start_row, end_row_exclusive):
        dest_r = r - start_row + 1
        for c in range(min_col, max_col + 1):
            src_c = c - min_col + 1
            src_cell = src_ws.cell(row=r, column=c)
            cell_value = _get_cell_value_respecting_merge(src_ws, r, c, merged_ranges)
            dest_cell = dest_ws.cell(row=dest_r, column=src_c, value=cell_value)
            _copy_cell_style(src_cell, dest_cell)
        # Copy row height
        if r in src_ws.row_dimensions and src_ws.row_dimensions[r].height is not None:
            dest_ws.row_dimensions[dest_r].height = src_ws.row_dimensions[r].height
    for c in range(min_col, max_col + 1):
        src_letter = get_column_letter(c)
        dest_c = c - min_col + 1
        dest_letter = get_column_letter(dest_c)
        if src_letter in src_ws.column_dimensions:
            dest_ws.column_dimensions[dest_letter].width = src_ws.column_dimensions[src_letter].width
    return wb


TDR_HEADER_MAP = [
    ("no_of_ban", ("number of ban", "no. of ban", "no of ban", "bans needed", "number of bans needed", "# of bans", "# of ban")),
    ("account_type", ("account type", "account segment")),
    ("sub_type", ("sub type", "subtype", "account subtype")),
    ("source_plan", ("source price plan", "source protection", "uscc plan")),
    ("line_type", ("line type", "t-mobile")),
    ("target_soc", ("target plan soc", "target plan boc", "target feature code")),
    ("target_plan_name", ("target plan name")),
    ("no_of_lines", ("no of lines")),
    ("owner", ("owner")),
    ("comment", ("comment")),
    ("bans_cell", ("od", "cid", "bans")),
]


def _find_tdr_section_header_row(ws, start_row, end_row):
    for r in range(start_row, end_row):
        row_tuple = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None)
        if not row_tuple:
            continue
        row_vals = list(row_tuple)
        no_ban_keywords = ("number of ban", "no. of ban", "no of ban", "bans needed", "number of bans needed", "# of bans", "# of ban")
        has_no_ban = _find_column_index_in_row(row_vals, no_ban_keywords) is not None
        has_account = _find_column_index_in_row(row_vals, ("account type", "account segment", "account subtype")) is not None
        if not (has_no_ban and has_account):
            continue
        col_map = []
        for field_name, keywords in TDR_HEADER_MAP:
            idx = _find_column_index_in_row(row_vals, keywords)
            if idx is not None:
                col_map.append((field_name, idx))
        if any(f[0] == "no_of_ban" for f in col_map) and any(f[0] == "bans_cell" for f in col_map):
            return (r, col_map)
    return (None, [])


def _extract_bans_list_from_cell(value):
    if value is None:
        return []
    return list(BAN_PATTERN.findall(str(value).strip()))


def _get_merged_ranges(ws):
    """Return list of (min_row, min_col, max_row, max_col) for each merged range. Requires ws not in read_only mode."""
    try:
        return [
            (r.min_row, r.min_col, r.max_row, r.max_col)
            for r in ws.merged_cells.ranges
        ]
    except Exception:
        return []


def _get_cell_value_respecting_merge(ws, row, col, merged_ranges):
    """Return cell value; if (row, col) is inside a merged range, return the top-left cell value."""
    for (min_row, min_col, max_row, max_col) in merged_ranges:
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return ws.cell(row=min_row, column=min_col).value
    return ws.cell(row=row, column=col).value


def extract_tdr_sections_with_rows(ws):
    tdr_rows = []
    for row in ws.iter_rows():
        row_idx = next((c.row for c in row if hasattr(c, "row")), None)
        if not row_idx:
            continue
        for cell in row:
            tdr_id = _extract_tdr_from_cell(cell.value)
            if tdr_id:
                tdr_rows.append((row_idx, tdr_id))
                break
    if not tdr_rows:
        return {}
    merged_ranges = _get_merged_ranges(ws)
    result = {}
    for i, (start_row, tdr_id) in enumerate(tdr_rows):
        end_row = tdr_rows[i + 1][0] if i + 1 < len(tdr_rows) else (ws.max_row + 1)
        header_row_idx, col_map = _find_tdr_section_header_row(ws, start_row, end_row)
        if header_row_idx is None or not col_map:
            continue
        col_by_field = dict(col_map)
        data_start = header_row_idx + 1
        rows_out = []
        for r in range(data_start, end_row):
            row_dict = {}
            for field_name, col_idx in col_by_field.items():
                val = _get_cell_value_respecting_merge(ws, r, col_idx, merged_ranges)
                if field_name == "bans_cell":
                    row_dict["bans_list"] = _extract_bans_list_from_cell(val)
                else:
                    row_dict[field_name] = val
            if "bans_list" not in row_dict:
                row_dict["bans_list"] = []
            no_ban = row_dict.get("no_of_ban")
            if no_ban is not None:
                try:
                    row_dict["no_of_ban"] = int(no_ban)
                except (TypeError, ValueError):
                    row_dict["no_of_ban"] = no_ban
            rows_out.append(row_dict)
        if rows_out:
            result[tdr_id] = rows_out
    return result


def _copy_sheet_into_workbook(src_ws, dest_wb, sheet_name):
    """Copy all cell values from src_ws into a new sheet in dest_wb with the given name."""
    dest_ws = dest_wb.create_sheet(title=sheet_name)
    for row in src_ws.iter_rows(values_only=True):
        dest_ws.append(row)
    return dest_ws


# Device Details sheet: columns to store as text to avoid long-number truncation/scientific notation
DEVICE_DETAILS_SHEET_NAME = "Pre-load device details"
BML_SHEET_NAME = "BML"
DEVICE_DETAILS_CUSTOMER_ID_HEADERS = ("customer_id", "customer id", "ban", "bans", "lgc_customer id")
DEVICE_DETAILS_TEXT_COLUMNS = (
    "customer_id", "customer id", "msisdn", "imei", "esn", "eid", "uiccid", "uimsi", "timsi",
    "device_model", "device_lock_status",
)


def _load_device_details(path, sheet_name=None):
    """
    Load device details Excel. Returns (headers, list of row lists, customer_id_col_1based) or None.
    Uses first sheet with CUSTOMER_ID (or similar) in header if sheet_name not given.
    """
    if not path or not os.path.isfile(path):
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet_to_use = sheet_name
        if not sheet_to_use:
            for name in wb.sheetnames:
                ws = wb[name]
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if first_row and _find_column_index_in_row(list(first_row), DEVICE_DETAILS_CUSTOMER_ID_HEADERS):
                    sheet_to_use = name
                    break
        if not sheet_to_use:
            wb.close()
            return None
        ws = wb[sheet_to_use]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return None
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        cust_col = _find_column_index_in_row(headers, DEVICE_DETAILS_CUSTOMER_ID_HEADERS)
        if cust_col is None:
            cust_col = 1
        return (headers, rows[1:], cust_col)
    except Exception:
        return None


def _add_device_details_sheet_to_workbook(wb, device_data, bans_set):
    """
    Add sheet 'Device Details' to wb with rows from device_data where CUSTOMER_ID is in bans_set.
    device_data = (headers, list of row tuples, customer_id_col_1based).
    Format long-number columns as text (@) to avoid scattering/truncation.
    """
    if not device_data or not bans_set:
        return
    headers, rows, cust_col = device_data
    cust_idx = cust_col - 1
    # Normalize BANs for match
    bans_normalized = set()
    for b in bans_set:
        n = _normalize_ban(b)
        if n:
            bans_normalized.add(n)
        if isinstance(b, str):
            bans_normalized.add(b.strip())
        else:
            bans_normalized.add(str(b).strip())
    matching = []
    for row in rows:
        if not row or len(row) <= cust_idx:
            continue
        val = row[cust_idx]
        n = _normalize_ban(val)
        if n and n in bans_normalized:
            matching.append(row)
            continue
        s = (str(val).strip() if val is not None else "")
        if s in bans_normalized:
            matching.append(row)
    # Sort by CUSTOMER_ID (column at cust_idx) ascending
    def _cust_sort_key(r):
        v = r[cust_idx] if r and len(r) > cust_idx else None
        if v is None:
            return ""
        s = str(v).strip()
        return s.zfill(20) if s.isdigit() else s
    matching.sort(key=_cust_sort_key)
    # Create sheet
    ws = wb.create_sheet(title=DEVICE_DETAILS_SHEET_NAME[:31])
    header_row = list(headers)
    ws.append(header_row)
    for row in matching:
        # Write as strings for long-number columns to avoid Excel reformatting
        out = []
        for i, v in enumerate(row):
            if i >= len(header_row):
                out.append(v)
                continue
            h = (header_row[i] or "").lower().replace(" ", "_")
            if any(t in h for t in ("customer_id", "msisdn", "imei", "esn", "eid", "uiccid", "uimsi", "timsi")):
                out.append(str(v) if v is not None else "")
            else:
                out.append(v)
        ws.append(out)
    # Set text format for long-number columns
    for col_idx, h in enumerate(header_row, 1):
        hl = (h or "").lower().replace(" ", "_")
        if any(t in hl for t in ("customer_id", "msisdn", "imei", "esn", "eid", "uiccid", "uimsi", "timsi")):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = "@"
    # Header styling
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(header_row) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.border = thin_border
    for r in range(2, ws.max_row + 1):
        for col in range(1, len(header_row) + 1):
            ws.cell(row=r, column=col).border = thin_border


def _add_bml_sheet_to_workbook(wb, bml_path, bans_set=None):
    """
    Add a sheet 'BML' to wb. If bans_set is provided, copy only rows where BAN/customer ID is in that set.
    Otherwise copy the full first sheet (backward compatible).
    """
    if not bml_path or not os.path.isfile(bml_path):
        return
    try:
        bml_wb = load_workbook(bml_path, read_only=False, data_only=True)
        if not bml_wb.sheetnames:
            bml_wb.close()
            return
        src_ws = bml_wb[bml_wb.sheetnames[0]]
        max_col = src_ws.max_column
        max_row = src_ws.max_row
        # Find BAN/customer ID column (1-based)
        header_row = next(src_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        ban_col = None
        if header_row:
            header_vals = list(header_row)
            ban_col = _find_column_index_in_row(header_vals, ("ban", "bans", "customer", "customer id", "account", "cid"))
        if not ban_col:
            ban_col = 1
        # Build set of normalized BANs to keep
        bans_normalized = set()
        if bans_set:
            for b in bans_set:
                n = _normalize_ban(b)
                if n:
                    bans_normalized.add(n)
                if isinstance(b, str) and b.strip():
                    bans_normalized.add(b.strip())
        dest_ws = wb.create_sheet(title=BML_SHEET_NAME[:31])
        # Copy header row
        for c in range(1, max_col + 1):
            dest_ws.cell(row=1, column=c, value=src_ws.cell(row=1, column=c).value)
        dest_row = 2
        for r in range(2, max_row + 1):
            if bans_normalized:
                ban_val = src_ws.cell(row=r, column=ban_col).value
                ban_n = _normalize_ban(ban_val)
                if not ban_n and ban_val is not None:
                    ban_n = str(ban_val).strip()
                if ban_n not in bans_normalized:
                    continue
            for c in range(1, max_col + 1):
                dest_ws.cell(row=dest_row, column=c, value=src_ws.cell(row=r, column=c).value)
            dest_row += 1
        for c in range(1, max_col + 1):
            letter = get_column_letter(c)
            if letter in src_ws.column_dimensions:
                dest_ws.column_dimensions[letter].width = src_ws.column_dimensions[letter].width
        bml_wb.close()
    except Exception:
        pass


def _copy_full_sheet_to_workbook(src_ws, dest_wb, sheet_title):
    """Copy all rows from src_ws into a new sheet in dest_wb with the given title (values only)."""
    dest_ws = dest_wb.create_sheet(title=(sheet_title[:31] if sheet_title else "Sheet"))
    for row in src_ws.iter_rows(values_only=True):
        dest_ws.append(list(row) if row else [])
    for c in range(1, src_ws.max_column + 1):
        letter = get_column_letter(c)
        if letter in src_ws.column_dimensions and src_ws.column_dimensions[letter].width is not None:
            dest_ws.column_dimensions[letter].width = src_ws.column_dimensions[letter].width


def _copy_full_sheet_to_workbook_with_format(src_ws, dest_wb, sheet_title):
    """Copy all rows and cell formatting (font, fill, border, alignment, number_format) from src_ws to a new sheet in dest_wb."""
    dest_ws = dest_wb.create_sheet(title=(sheet_title[:31] if sheet_title else "Sheet"))
    for row in src_ws.iter_rows():
        row_idx = row[0].row if row and hasattr(row[0], "row") else None
        if not row_idx:
            continue
        dest_row_idx = row_idx
        for cell in row:
            dest_cell = dest_ws.cell(row=dest_row_idx, column=cell.column, value=cell.value)
            _copy_cell_style(cell, dest_cell)
        if row_idx in src_ws.row_dimensions and src_ws.row_dimensions[row_idx].height is not None:
            dest_ws.row_dimensions[dest_row_idx].height = src_ws.row_dimensions[row_idx].height
    for c in range(1, src_ws.max_column + 1):
        letter = get_column_letter(c)
        if letter in src_ws.column_dimensions and src_ws.column_dimensions[letter].width is not None:
            dest_ws.column_dimensions[letter].width = src_ws.column_dimensions[letter].width


# QE_MBL sheet formatting (match user's example: font size, color, background)
# Header: dark teal background, white bold text (for BML and Device Details)
QE_MBL_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
QE_MBL_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
# Alternating row colors
QE_MBL_BML_ALT_COLOR = "DDEBF7"       # light blue/aqua for odd data rows
QE_MBL_DEVICE_ALT_COLOR = "E2EFDA"    # light green for odd data rows
QE_MBL_DATA_FONT = Font(color="000000", size=11)
QE_MBL_THIN_BORDER = Border(
    left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"),
)
# Delivery Status: magenta header, white text; first data row grey; D/E/F center; G wrap
QE_MBL_DELIVERY_STATUS_HEADER_FILL = PatternFill(start_color="FF009F", end_color="FF009F", fill_type="solid")
QE_MBL_DELIVERY_STATUS_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
QE_MBL_DELIVERY_STATUS_ROW2_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
QE_MBL_DELIVERY_STATUS_COLUMN_WIDTHS = (20.57, 12.71, 19.71, 20.0, 24.43, 23.29, 88.14)


def _qe_mbl_sheet_dimensions(ws):
    """Return (max_row, max_col) for sheet, computing from cells if dimensions are not set."""
    max_r, max_c = getattr(ws, "max_row", 0) or 0, getattr(ws, "max_column", 0) or 0
    if max_r > 0 and max_c > 0:
        return max_r, max_c
    for row in ws.iter_rows():
        for cell in row:
            if cell.row and cell.column:
                max_r = max(max_r, cell.row)
                max_c = max(max_c, cell.column)
    return max_r, max_c


def _apply_qe_mbl_sheet_format(ws, alt_fill_color, alt_on_even_row=True):
    """Apply QE_MBL format: row 1 dark teal header (white bold); data rows alternating alt_fill_color and white.
    alt_on_even_row: if True, rows 2,4,6 get alt color (BML); if False, rows 3,5,7 get alt color (Device Details)."""
    max_row, max_col = _qe_mbl_sheet_dimensions(ws)
    if max_row < 1 or max_col < 1:
        return
    # Row 1 header: give each cell its own style copy so Excel persists formatting
    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col)
        c.fill = copy(QE_MBL_HEADER_FILL)
        c.font = copy(QE_MBL_HEADER_FONT)
        c.border = copy(QE_MBL_THIN_BORDER)
    for r in range(2, max_row + 1):
        use_alt = (r % 2 == 0) if alt_on_even_row else (r % 2 == 1)
        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = PatternFill(start_color=alt_fill_color, end_color=alt_fill_color, fill_type="solid") if use_alt else PatternFill()
            cell.font = copy(QE_MBL_DATA_FONT)
            cell.border = copy(QE_MBL_THIN_BORDER)


def _format_delivery_status_sheet(ws):
    """Apply formatting: magenta header with white bold text; row 2 grey; D/E/F center; G wrap; thin borders."""
    max_row = getattr(ws, "max_row", 0) or 0
    if max_row < 1:
        return
    # Header row 1
    for col in range(1, 8):
        c = ws.cell(row=1, column=col)
        c.fill = copy(QE_MBL_DELIVERY_STATUS_HEADER_FILL)
        c.font = copy(QE_MBL_DELIVERY_STATUS_HEADER_FONT)
        c.border = copy(QE_MBL_THIN_BORDER)
    for i, width in enumerate(QE_MBL_DELIVERY_STATUS_COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for r in range(2, max_row + 1):
        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            cell.border = copy(QE_MBL_THIN_BORDER)
            cell.font = copy(QE_MBL_DATA_FONT)
            if col == 7:
                cell.alignment = copy(wrap_align)
            elif col in (4, 5, 6):
                cell.alignment = copy(center_align)
            if r == 2:
                cell.fill = copy(QE_MBL_DELIVERY_STATUS_ROW2_FILL)
            elif r > 2:
                cell.fill = PatternFill()


def build_qe_mbl_ban_list_workbook(bml_path, device_details_path, delivery_status_rows, device_details_sheet_name=None):
    """
    Build QE_MBL_BAN_LIST workbook with 3 sheets: BML (full), Pre-load device details (full), Delivery Status.
    Copies formatting from source sheets for BML and Device Details; applies example formatting to Delivery Status.
    delivery_status_rows: list of (tdr_id, requestor, status, asked, delivered, dfs_load, comment).
    Returns workbook bytes.
    """
    from io import BytesIO
    wb = Workbook()
    wb.remove(wb.active)
    if bml_path and os.path.isfile(bml_path):
        try:
            bml_wb = load_workbook(bml_path, read_only=False, data_only=False)
            if bml_wb.sheetnames:
                _copy_full_sheet_to_workbook_with_format(bml_wb[bml_wb.sheetnames[0]], wb, "BML")
            bml_wb.close()
        except Exception:
            try:
                bml_wb = load_workbook(bml_path, read_only=False, data_only=True)
                if bml_wb.sheetnames:
                    _copy_full_sheet_to_workbook(bml_wb[bml_wb.sheetnames[0]], wb, "BML")
                bml_wb.close()
            except Exception:
                pass
        if "BML" in wb.sheetnames:
            _apply_qe_mbl_sheet_format(wb["BML"], QE_MBL_BML_ALT_COLOR, alt_on_even_row=True)
    if device_details_path and os.path.isfile(device_details_path):
        try:
            dev_wb = load_workbook(device_details_path, read_only=False, data_only=False)
            sheet_name = device_details_sheet_name or DEVICE_DETAILS_SHEET_NAME
            if sheet_name in dev_wb.sheetnames:
                _copy_full_sheet_to_workbook_with_format(dev_wb[sheet_name], wb, DEVICE_DETAILS_SHEET_NAME)
            elif dev_wb.sheetnames:
                _copy_full_sheet_to_workbook_with_format(dev_wb[dev_wb.sheetnames[0]], wb, DEVICE_DETAILS_SHEET_NAME)
            dev_wb.close()
        except Exception:
            try:
                dev_wb = load_workbook(device_details_path, read_only=False, data_only=True)
                sheet_name = device_details_sheet_name or DEVICE_DETAILS_SHEET_NAME
                if sheet_name in dev_wb.sheetnames:
                    _copy_full_sheet_to_workbook(dev_wb[sheet_name], wb, DEVICE_DETAILS_SHEET_NAME)
                elif dev_wb.sheetnames:
                    _copy_full_sheet_to_workbook(dev_wb[dev_wb.sheetnames[0]], wb, DEVICE_DETAILS_SHEET_NAME)
                dev_wb.close()
            except Exception:
                pass
        if DEVICE_DETAILS_SHEET_NAME in wb.sheetnames:
            _apply_qe_mbl_sheet_format(wb[DEVICE_DETAILS_SHEET_NAME], QE_MBL_DEVICE_ALT_COLOR, alt_on_even_row=False)
    headers = ["TDR ID", "Requestor", "Status", "# of BANs Asked", "# of BANs Delivered", "DFS Load Required", "Comment"]
    ws_ds = wb.create_sheet(title="Delivery Status")
    ws_ds.append(headers)
    for row_tuple in (delivery_status_rows or []):
        ws_ds.append(list(row_tuple))
    _format_delivery_status_sheet(ws_ds)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _normalize_ban(value):
    """Return 9-digit BAN as string, or None if not a valid BAN."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        v = int(value)
        if 100_000_000 <= v <= 999_999_999:
            return str(v)
        return None
    s = str(value).strip()
    if s.isdigit() and len(s) == 9:
        return s
    m = BAN_PATTERN.search(s)
    return m.group(1) if m else None


def _find_column_index_in_row(row_values, header_keywords):
    """Return 1-based column index of first cell in row that contains any keyword (case-insensitive)."""
    keywords = [k.lower() for k in header_keywords]
    for col_idx, cell_value in enumerate(row_values or (), 1):
        if cell_value is None:
            continue
        val = str(cell_value).strip().lower()
        if any(kw in val for kw in keywords):
            return col_idx
    return None


def _find_header_row(ws, max_look=10):
    """Find first row that has both BAN-like and Status-like headers. Returns (1-based row index, row values) or (None, None)."""
    ban_kw = ("ban", "bans", "ban id", "account")
    status_kw = ("status", "result", "lvt", "passed", "delivered", "outcome", "state", "execution")
    for r in range(1, min(ws.max_row + 1, max_look + 1)):
        row_tuple = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None)
        if not row_tuple:
            continue
        row_vals = list(row_tuple)
        has_ban = _find_column_index_in_row(row_vals, ban_kw) is not None
        has_status = _find_column_index_in_row(row_vals, status_kw) is not None
        if has_ban and has_status:
            return (r, row_vals)
    return (None, None)


def _find_ban_column_in_lvt(ws):
    """Return 1-based column index for BAN/customer ID. Match if header contains any keyword."""
    ban_keywords = ("ban", "bans", "customer", "account", "cid", "lgc")
    for r in range(1, min(ws.max_row + 1, 15)):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None)
        if not row:
            continue
        for c, val in enumerate(row):
            if val is None:
                continue
            v = str(val).strip().lower()
            if any(kw in v for kw in ban_keywords) and "id" in v or v in ("ban", "bans", "account", "cid"):
                return c + 1
    return 1


def _find_status_column_in_lvt(ws):
    """Return 1-based column index for Status. Match if header contains any keyword."""
    status_keywords = ("status", "result", "lvt", "pass", "fail", "verified", "outcome")
    for r in range(1, min(ws.max_row + 1, 15)):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None)
        if not row:
            continue
        for c, val in enumerate(row):
            if val is None:
                continue
            v = str(val).strip().lower()
            if any(kw in v for kw in status_keywords):
                return c + 1
    return 2


def _build_ban_to_status_from_sheet(ws):
    """Build dict BAN (str) -> Status from LVT sheet. Same logic as bulk: all rows from row 2 with Passed/Failed."""
    ban_to_status = {}
    ban_col = _find_ban_column_in_lvt(ws)
    status_col = _find_status_column_in_lvt(ws)
    ban_header_keywords = (
        "ban", "bans", "customer", "customer id", "account", "cid",
        "lgc_customer_id", "customer_id", "status", "lgc customer id",
    )
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or ban_col > len(row):
            continue
        val = row[ban_col - 1]
        if val is None:
            continue
        cid = str(val).strip()
        if not cid or cid.lower() in ban_header_keywords:
            continue
        if "customer" in cid.lower() and "id" in cid.lower():
            continue
        if cid.replace("_", "").replace("-", "").replace(" ", "").isalpha():
            continue
        status_val = row[status_col - 1] if status_col <= len(row) else None
        status_str = (str(status_val).strip() if status_val is not None else "").strip()
        key = _normalize_ban(val) or cid
        ban_to_status[key] = status_str
    return ban_to_status


def _fill_tdr_info_status_column(tdr_ws, ban_to_status):
    """Fill column C (Status) in TDR Info sheet by looking up BAN in ban_to_status. Uses 'Not found' when BAN not in BAN Wise Result."""
    tdr_ws.cell(row=1, column=3, value="Status")
    filled = 0
    for r in range(2, tdr_ws.max_row + 1):
        ban_cell = tdr_ws.cell(row=r, column=2).value
        ban_str = _normalize_ban(ban_cell)
        status = ban_to_status.get(ban_str, "Not found") if ban_str else "Not found"
        if status and status != "Not found":
            filled += 1
        tdr_ws.cell(row=r, column=3, value=status)
    return filled


def _find_failures_header_row(ws, max_look=10):
    """Find row with customer/BAN, description, and check_id. Returns (1-based row index, row values) or (None, None)."""
    cust_kw = ("lgc_customer", "customer id", "customer", "ban", "bans", "account")
    desc_kw = ("description", "failure", "desc")
    check_kw = ("check_id", "check id", "checkid")
    for r in range(1, min(ws.max_row + 1, max_look + 1)):
        row_tuple = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None)
        if not row_tuple:
            continue
        row_vals = list(row_tuple)
        if (_find_column_index_in_row(row_vals, cust_kw) is not None and
                _find_column_index_in_row(row_vals, desc_kw) is not None and
                _find_column_index_in_row(row_vals, check_kw) is not None):
            return (r, row_vals)
    return (None, None)


def _build_ban_to_failures_from_sheet(ws):
    """Build dict BAN (str) -> list of (description, check_id) from BAN Wise Failures sheet."""
    from collections import defaultdict
    ban_to_failures = defaultdict(list)

    cust_kw = ("lgc_customer", "customer id", "customer", "ban", "bans", "account")
    desc_kw = ("description", "failure", "desc")
    check_kw = ("check_id", "check id", "checkid")

    header_row_idx, header_row_vals = _find_failures_header_row(ws)
    if header_row_idx is None or not header_row_vals:
        return dict(ban_to_failures)

    cust_col = _find_column_index_in_row(header_row_vals, cust_kw)
    desc_col = _find_column_index_in_row(header_row_vals, desc_kw)
    check_col = _find_column_index_in_row(header_row_vals, check_kw)
    if not all((cust_col, desc_col, check_col)):
        return dict(ban_to_failures)

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or len(row) < max(cust_col, desc_col, check_col):
            continue
        cust_val = row[cust_col - 1]
        desc_val = row[desc_col - 1]
        check_val = row[check_col - 1]
        ban_strs = set()
        if cust_val is not None:
            s = str(cust_val).strip()
            for m in BAN_PATTERN.finditer(s):
                ban_strs.add(m.group(1))
            if not ban_strs:
                single = _normalize_ban(cust_val)
                if single:
                    ban_strs.add(single)
        desc_str = (str(desc_val).strip() if desc_val is not None else "") or ""
        if check_val is None:
            check_str = ""
        else:
            s = str(check_val).strip()
            check_str = str(int(float(check_val))) if s.replace(".", "").isdigit() else s
        for ban_str in ban_strs:
            ban_to_failures[ban_str].append((desc_str, check_str))

    return dict(ban_to_failures)


def _fill_tdr_info_failure_columns(tdr_ws, ban_to_failures):
    """Fill columns D (Failure Description), E (Check ID), F (Comments). For Passed status use N/A; else look up BAN in ban_to_failures."""
    tdr_ws.cell(row=1, column=4, value="Failure Description")
    tdr_ws.cell(row=1, column=5, value="Check ID")
    tdr_ws.cell(row=1, column=6, value="Comments")
    for r in range(2, tdr_ws.max_row + 1):
        status_cell = tdr_ws.cell(row=r, column=3).value
        status_str = (str(status_cell).strip().lower() if status_cell else "") or ""
        if status_str == "passed":
            tdr_ws.cell(row=r, column=4, value="N/A")
            tdr_ws.cell(row=r, column=5, value="N/A")
            tdr_ws.cell(row=r, column=6, value="")
            continue
        ban_cell = tdr_ws.cell(row=r, column=2).value
        ban_str = _normalize_ban(ban_cell)
        if not ban_str or ban_str not in ban_to_failures:
            tdr_ws.cell(row=r, column=4, value="Not found")
            tdr_ws.cell(row=r, column=5, value="Not found")
            tdr_ws.cell(row=r, column=6, value="")
            continue
        pairs = ban_to_failures[ban_str]
        descriptions = [p[0] for p in pairs if p[0]]
        check_ids = [p[1] for p in pairs if p[1]]
        tdr_ws.cell(row=r, column=4, value="; ".join(descriptions) if descriptions else "Not found")
        tdr_ws.cell(row=r, column=5, value=", ".join(check_ids) if check_ids else "Not found")
        # Column F: comments for known check IDs (e.g. 3106 -> "BANs can be used after RS-122932 fix.")
        comments = []
        for cid in (check_ids or []):
            cid_str = str(cid).strip()
            if cid_str and cid_str in CHECK_ID_COMMENTS:
                comments.append(CHECK_ID_COMMENTS[cid_str])
        tdr_ws.cell(row=r, column=6, value="; ".join(comments) if comments else "")


def _format_tdr_info_sheet(ws):
    """Apply formatting to TDR Info sheet: header style, column widths, status colors, borders."""
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    passed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    notfound_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    num_cols = max(6, ws.max_column)
    for col in range(1, num_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 36
    for r in range(2, ws.max_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 3:
                val = (cell.value or "").strip().lower()
                if val == "passed":
                    cell.fill = passed_fill
                elif val == "failed":
                    cell.fill = failed_fill
                elif val == "not found":
                    cell.fill = notfound_fill
    ws.freeze_panes = "A2"


def _build_tdr_summary(all_rows, ban_to_status):
    """Return list of (TDR, total, passed, failed, not_found, tdr_status). TDR status: Passed, Failed, or Partial (some not found/OOS)."""
    from collections import defaultdict
    counts = defaultdict(lambda: {"passed": 0, "failed": 0, "not_found": 0})
    for tdr, ban in all_rows:
        ban_str = _normalize_ban(ban)
        status = ban_to_status.get(ban_str, "Not found") if ban_str else "Not found"
        s = str(status).strip().lower()
        if status == "Not found" or s == "not found":
            counts[tdr]["not_found"] += 1
        elif s == "passed":
            counts[tdr]["passed"] += 1
        elif s == "failed":
            counts[tdr]["failed"] += 1
    result = []
    for tdr in sorted(counts.keys()):
        c = counts[tdr]
        total = c["passed"] + c["failed"] + c["not_found"]
        tdr_status = "Failed" if c["failed"] > 0 else ("Partial" if c["not_found"] > 0 else "Passed")
        result.append((tdr, total, c["passed"], c["failed"], c["not_found"], tdr_status))
    return result


def _add_mapping_sheet(wb, ban_to_status, ban_to_source):
    """Add 'Mapping' sheet: all LVT customer IDs with TDR Number, Status, LVT Status, Excel File, Sheet Name (like bulk TDR mapping)."""
    ws = wb.active
    ws.title = "Mapping"
    headers = ["Customer ID", "TDR Number", "Status", "LVT Status", "Excel File", "Sheet Name"]
    ws.append(headers)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col, value=headers[col - 1])
        c.fill = header_fill
        c.font = header_font
        c.border = thin_border
    # One row per LVT customer; sort: Found first, then Not found; by TDR then Customer ID
    _status_order = {"Found": 0, "Not found": 1}
    rows = []
    for ban_str in ban_to_status:
        src = ban_to_source.get(ban_str) or ban_to_source.get(_normalize_ban(ban_str) if ban_str else "")
        if src:
            tdr_id, excel_file, sheet_name = src
            status = "Found"
        else:
            tdr_id, excel_file, sheet_name = "", "", ""
            status = "Not found"
        lvt_status = (ban_to_status.get(ban_str) or "").strip() or ""
        rows.append((_status_order.get(status, 1), tdr_id or "ZZZ", ban_str, ban_str, tdr_id, status, lvt_status, excel_file, sheet_name))
    rows.sort(key=lambda r: (r[0], r[1], r[3]))
    for _ord, _tdr_sort, _cid_sort, cid, tdr_num, status, lvt_st, excel_file, sheet_name in rows:
        ws.append([cid, tdr_num, status, lvt_st, excel_file, sheet_name])
    for r in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = thin_border
    for i, w in enumerate([16, 16, 12, 12, 20, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 50)
    ws.freeze_panes = "A2"


def _add_tdr_summary_sheet(wb, tdr_summary_rows):
    """Add 'TDR Summary' sheet with TDR-wise counts and status."""
    ws = wb.create_sheet(title="TDR Summary")
    ws.append(["TDR", "Total BANs", "Passed", "Failed", "Not found", "TDR Status"])
    for row in tdr_summary_rows:
        ws.append(list(row))
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    passed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    for col in range(1, 7):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border
    for i, w in enumerate([18, 12, 10, 10, 12, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.border = thin_border
            if col == 6:
                val = (cell.value or "").strip()
                if val == "Passed":
                    cell.fill = passed_fill
                elif val == "Failed":
                    cell.fill = failed_fill
                elif val == "Partial":
                    cell.fill = partial_fill
    ws.freeze_panes = "A2"


def _safe_tdr_filename(tdr_id):
    """Return a safe filename for the TDR (e.g. TDR-204119 -> TDR-204119.xlsx)."""
    safe = re.sub(r'[\\/:*?"<>|]', "_", str(tdr_id).strip())
    return f"{safe}.xlsx" if safe else "TDR.xlsx"


def _row_val(v):
    if v is None:
        return ""
    return str(v).strip() if v != "" else ""


def _write_one_excel_per_tdr(all_rows, ban_to_status, output_folder, tdr_sections_data=None):
    from collections import defaultdict
    os.makedirs(output_folder, exist_ok=True)
    written = []
    all_tdrs = sorted(set(t for t, _ in all_rows))
    if tdr_sections_data:
        out_headers = ["SCN No", "NO. of BAN", "Account Type", "Sub Type", "USCC plan", "T-Mobile Plan SOC &Name", "USCC Plan", "Target Plan SOC &Name", "No of Lines", "BANs", "Comment"]
        for tdr in all_tdrs:
            if tdr not in tdr_sections_data:
                continue
            section_rows = tdr_sections_data[tdr]
            wb = Workbook()
            ws = wb.active
            ws.title = tdr[:31]
            ws.append(out_headers)
            for i, row_dict in enumerate(section_rows, 1):
                scn = f"SCN{i:02d}"
                no_ban = row_dict.get("no_of_ban", "")
                bans_list = row_dict.get("bans_list") or []
                provided_count = len(bans_list)
                needed = no_ban if isinstance(no_ban, int) else (int(no_ban) if str(no_ban).strip().isdigit() else 0)
                comment = _row_val(row_dict.get("comment"))
                if provided_count >= needed or needed == 0:
                    comment = ""
                bans_str = "\n".join(bans_list) if bans_list else ("N/A" if needed else "0")
                if not bans_str and needed:
                    bans_str = "N/A"
                ws.append([scn, no_ban, _row_val(row_dict.get("account_type")), _row_val(row_dict.get("sub_type")), _row_val(row_dict.get("source_plan")), _row_val(row_dict.get("line_type")), _row_val(row_dict.get("target_soc")), _row_val(row_dict.get("target_plan_name")), _row_val(row_dict.get("no_of_lines")), bans_str, comment])
            _format_tdr_per_sheet_wide(ws)
            path = os.path.join(output_folder, _safe_tdr_filename(tdr))
            try:
                wb.save(path)
                written.append(path)
            except PermissionError:
                pass
        for tdr in all_tdrs:
            if tdr in tdr_sections_data:
                continue
            by_tdr = defaultdict(list)
            for t, ban in all_rows:
                if t != tdr:
                    continue
                ban_str = _normalize_ban(ban)
                status = ban_to_status.get(ban_str, "Not found") if ban_str else "Not found"
                by_tdr[tdr].append((tdr, ban, status))
            if not by_tdr.get(tdr):
                continue
            wb = Workbook()
            ws = wb.active
            ws.title = tdr[:31]
            ws.append(["TDR", "BAN", "Status"])
            for r in by_tdr[tdr]:
                ws.append(list(r))
            _format_tdr_info_sheet(ws)
            path = os.path.join(output_folder, _safe_tdr_filename(tdr))
            try:
                wb.save(path)
                written.append(path)
            except PermissionError:
                pass
        return written
    by_tdr = defaultdict(list)
    for tdr, ban in all_rows:
        ban_str = _normalize_ban(ban)
        status = ban_to_status.get(ban_str, "Not found") if ban_str else "Not found"
        by_tdr[tdr].append((tdr, ban, status))
    for tdr in sorted(by_tdr.keys()):
        rows = by_tdr[tdr]
        wb = Workbook()
        ws = wb.active
        ws.title = tdr[:31]
        ws.append(["TDR", "BAN", "Status"])
        for r in rows:
            ws.append(list(r))
        _format_tdr_info_sheet(ws)
        path = os.path.join(output_folder, _safe_tdr_filename(tdr))
        try:
            wb.save(path)
            written.append(path)
        except PermissionError:
            pass
    return written


def _format_tdr_per_sheet_wide(ws):
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, 12):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border
    widths = [10, 12, 12, 10, 22, 22, 14, 24, 12, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 50)
    for r in range(2, ws.max_row + 1):
        for col in range(1, 12):
            ws.cell(row=r, column=col).border = thin_border
    ws.freeze_panes = "A2"


def run_extraction_and_report(all_sources, output_excel=None, lvt_report_path=None, lvt_sheet_name=None, device_details_path=None, device_details_sheet_name=None, bml_path=None, source_display_names=None):
    """
    Extract TDR→BAN from Data details; keep only BANs that exist in LVT.
    source_display_names: optional list of display names for each item in all_sources (e.g. original uploaded file names).
    """
    wb = None
    all_rows = []
    ban_to_source = {}  # normalized_ban -> (tdr_id, excel_filename, sheet_name) for Mapping sheet
    ban_to_source_path = {}  # normalized_ban -> (excel_path, sheet_name) so we copy TDR from the sheet that has the customer
    tdr_sections_data = {}
    tdr_section_ranges = []  # (tdr_id, excel_path, sheet_name, start_row, end_row_exclusive)
    tdr_excel_folder = os.path.join(REPORT_FOLDER, datetime.now().strftime("%Y%m%d") + "_TDR")
    per_tdr_files = set()
    os.makedirs(tdr_excel_folder, exist_ok=True)

    for i, (excel_path, sheet_names) in enumerate(all_sources):
        excel_display_name = (source_display_names[i] if source_display_names and i < len(source_display_names) else os.path.basename(excel_path))
        wb = load_workbook(excel_path, read_only=False, data_only=True)
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            mapping = extract_tdr_ban_mapping(ws)
            sections_rows = extract_tdr_sections_with_rows(ws)
            for tdr_id, rows in sections_rows.items():
                tdr_sections_data[tdr_id] = rows
            if not mapping:
                continue
            for tdr, bans in mapping.items():
                for ban in bans:
                    all_rows.append((tdr, ban))
                    nban = _normalize_ban(ban)
                    if nban and nban not in ban_to_source:
                        ban_to_source[nban] = (tdr, excel_display_name, sheet_name)
                        ban_to_source_path[nban] = (excel_path, sheet_name)
            for tdr_id, start_row, end_row_exclusive in get_tdr_section_ranges(ws):
                tdr_section_ranges.append((tdr_id, excel_path, sheet_name, start_row, end_row_exclusive))
        if wb:
            wb.close()
            wb = None

    # Load LVT and keep only (TDR, BAN) rows whose BAN is in LVT
    ban_to_status = {}
    ban_to_failures = {}
    sheet_to_use = (lvt_sheet_name or LVT_SHEET_NAME) if lvt_report_path else None
    if lvt_report_path and os.path.isfile(lvt_report_path) and sheet_to_use:
        try:
            lvt_wb = load_workbook(lvt_report_path, read_only=True, data_only=True)
            if sheet_to_use in lvt_wb.sheetnames:
                ban_to_status = _build_ban_to_status_from_sheet(lvt_wb[sheet_to_use])
            failures_sheet_name = _find_failures_sheet_in_workbook(lvt_wb)
            if failures_sheet_name is not None:
                ban_to_failures = _build_ban_to_failures_from_sheet(lvt_wb[failures_sheet_name])
            lvt_wb.close()
        except Exception:
            pass

    lvt_filter_applied = bool(ban_to_status)
    rows_in_lvt = [(t, b) for t, b in all_rows if _normalize_ban(b) in ban_to_status] if ban_to_status else all_rows

    # Per-TDR Excels: only for TDRs with at least one LVT-matched BAN; copy from the sheet where those BANs exist (same TDR can appear on multiple sheets)
    lvt_tdr_ids = sorted(set(t for t, _ in rows_in_lvt))
    seen_tdr = set()
    for tdr_id in lvt_tdr_ids:
        if tdr_id in seen_tdr:
            continue
        seen_tdr.add(tdr_id)
        bans_for_tdr = {ban for t, ban in rows_in_lvt if t == tdr_id}
        # Pick the (excel_path, sheet_name) where this TDR's LVT customers were found (so we copy from that sheet only)
        sheet_counts = {}
        for ban in bans_for_tdr:
            nban = _normalize_ban(ban)
            key = ban_to_source_path.get(nban)
            if key:
                sheet_counts[key] = sheet_counts.get(key, 0) + 1
        if not sheet_counts:
            continue
        excel_path, sheet_name = max(sheet_counts, key=sheet_counts.get)
        range_entry = next((r for r in tdr_section_ranges if r[0] == tdr_id and r[1] == excel_path and r[2] == sheet_name), None)
        if not range_entry:
            continue
        _tid, excel_path, sheet_name, start_row, end_row_exclusive = range_entry
        try:
            src_wb = load_workbook(excel_path, read_only=False, data_only=True)
            if sheet_name not in src_wb.sheetnames:
                src_wb.close()
                continue
            src_ws = src_wb[sheet_name]
            copy_wb = _copy_sheet_range_to_workbook(src_ws, start_row, end_row_exclusive, sheet_title=tdr_id)
            src_wb.close()
            if device_details_path and os.path.isfile(device_details_path):
                device_data = _load_device_details(device_details_path, device_details_sheet_name)
                if device_data:
                    _add_device_details_sheet_to_workbook(copy_wb, device_data, bans_for_tdr)
            if bml_path and os.path.isfile(bml_path):
                _add_bml_sheet_to_workbook(copy_wb, bml_path, bans_for_tdr)
            path = os.path.join(tdr_excel_folder, _safe_tdr_filename(tdr_id))
            copy_wb.save(path)
            per_tdr_files.add(path)
        except (PermissionError, Exception):
            pass

    # Delivery Status rows for QE_MBL_BAN_LIST: row 1 = summary, row 2+ = per TDR (TDR ID, QE Team, Status, # BANs Asked, # BANs Delivered, DFS No, Comment)
    DELIVERY_STATUS_DEFAULT_COMMENT = (
        'Note:\nIf delivered BANs are not part of "pre-load device details", meaning they were shared earlier '
        "and already converted and now mapped and delivered as part of this scenario"
    )
    delivery_status_rows = []
    total_asked = 0
    total_delivered = 0
    for tdr_id in lvt_tdr_ids:
        section_rows = tdr_sections_data.get(tdr_id, [])
        asked = 0
        for row_dict in section_rows:
            no_ban = row_dict.get("no_of_ban")
            if no_ban is not None:
                try:
                    asked += int(no_ban)
                except (TypeError, ValueError):
                    pass
        delivered = sum(1 for t, _ in rows_in_lvt if t == tdr_id)
        status = "Full Delivery" if asked == delivered else "Partial Delivery"
        delivery_status_rows.append((tdr_id, "QE Team", status, asked, delivered, "No", DELIVERY_STATUS_DEFAULT_COMMENT))
        total_asked += asked
        total_delivered += delivered
    # Always add summary row first (so QE_MBL download has at least one data row and third button shows)
    row1_status = "Full Delivery" if total_asked == total_delivered else "Partial Delivery"
    delivery_status_rows.insert(0, ("QE Sanity Data", "QE Team", row1_status, total_asked, total_delivered, "No", "NA"))

    # Single Excel: Mapping (all LVT customers) + TDR Info + TDR Summary
    if output_excel and ban_to_status:
        # Summary from all LVT customers (same as bulk mapping)
        summary = {"total": len(ban_to_status), "passed": 0, "failed": 0, "not_found": 0}
        for status in ban_to_status.values():
            st = str(status).strip().lower()
            if st == "passed":
                summary["passed"] += 1
            elif st == "failed":
                summary["failed"] += 1
            else:
                summary["not_found"] += 1

        out_wb = Workbook()
        if ban_to_status:
            _add_mapping_sheet(out_wb, ban_to_status, ban_to_source)
        out_ws = out_wb.create_sheet(title="TDR Info", index=1) if ban_to_status else out_wb.active
        if not ban_to_status:
            out_ws.title = "TDR Info"
        out_ws.append(["TDR", "BAN", "Status"])
        for row in rows_in_lvt:
            out_ws.append(list(row))
        _fill_tdr_info_status_column(out_ws, ban_to_status)
        _fill_tdr_info_failure_columns(out_ws, ban_to_failures)
        tdr_summary_rows = _build_tdr_summary(rows_in_lvt, ban_to_status)
        _add_tdr_summary_sheet(out_wb, tdr_summary_rows)
        summary["tdr_passed"] = sum(1 for _r in tdr_summary_rows if _r[5] == "Passed")
        summary["tdr_failed"] = sum(1 for _r in tdr_summary_rows if _r[5] == "Failed")
        summary["tdr_partial"] = sum(1 for _r in tdr_summary_rows if _r[5] == "Partial")
        _format_tdr_info_sheet(out_ws)
        summary["per_tdr_folder"] = tdr_excel_folder
        summary["per_tdr_count"] = len(per_tdr_files)
        summary["per_tdr_file_names"] = [os.path.basename(p) for p in per_tdr_files]
        summary["lvt_filter_applied"] = lvt_filter_applied
        summary["delivery_status_rows"] = delivery_status_rows
        try:
            out_wb.save(output_excel)
            return (output_excel, summary)
        except PermissionError:
            print(f"\n  Permission denied: {output_excel}")
            print("  Close the file if it is open in Excel, then run again.")
            return (None, None)
    return (None, None)


def _box_title(title):
    print(f"{C.CYAN}{C.BOLD}{BOX['TL']}{BOX['H'] * (W - 2)}{BOX['TR']}{C.RESET}")
    print(f"{C.CYAN}{BOX['V']}{C.RESET} {C.BOLD}{title:<{W - 4}}{C.RESET} {C.CYAN}{BOX['V']}{C.RESET}")
    print(f"{C.CYAN}{BOX['BL']}{BOX['H'] * (W - 2)}{BOX['BR']}{C.RESET}")


def main():
    _box_title("TDR Data Excel – Sheet & file selection")
    print(f"\n{C.DIM}Working folder:{C.RESET} {BASE_FOLDER}")

    # 1) TDR Data Excel path
    excel_path = get_excel_path()
    if not excel_path:
        return

    # 2) Which sheet(s) to use from this file
    sheet_names = get_sheet_names(excel_path)
    if not sheet_names:
        print("  No sheets found in workbook.")
        return

    selected_sheets = ask_sheet_choice(sheet_names, os.path.basename(excel_path))
    if not selected_sheets:
        print("  No sheet selected. Exiting.")
        return

    # 3) Optional: more Excel files or more sheets
    all_sources = [(excel_path, selected_sheets)]

    while ask_more_sources():
        more_path = input("  Path to another Excel file: ").strip()
        if not more_path or not os.path.isfile(more_path):
            print("  Invalid or missing file. Skipping.")
            continue
        more_sheets = get_sheet_names(more_path)
        if not more_sheets:
            print("  No sheets in that workbook. Skipping.")
            continue
        more_selected = ask_sheet_choice(more_sheets, os.path.basename(more_path))
        if more_selected:
            all_sources.append((more_path, more_selected))

    # 4) LVT report – user selects file from TDR deliver folder, then sheet if needed
    lvt_choice = get_lvt_report_file_and_sheet()
    lvt_report_path = lvt_choice[0] if lvt_choice else None
    lvt_sheet_name = lvt_choice[1] if lvt_choice else None

    # Extract TDR → BAN mapping and save single Excel to report folder
    os.makedirs(REPORT_FOLDER, exist_ok=True)
    archive_old_reports(REPORT_FOLDER, max_age_days=1)
    _ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_excel = os.path.join(REPORT_FOLDER, f"TDR_BAN_Report_{_ts}.xlsx")
    saved_path, summary = run_extraction_and_report(
        all_sources,
        output_excel=out_excel,
        lvt_report_path=lvt_report_path,
        lvt_sheet_name=lvt_sheet_name,
    )
    if saved_path:
        total = summary.get("total", 0) if summary else 0
        pct = lambda n: f"({100 * n / total:.1f}%)" if total else "(-)"
        print(f"\n{C.GREEN}{C.BOLD}{BOX['TL']}{BOX['H'] * (W - 2)}{BOX['TR']}{C.RESET}")
        print(f"{C.GREEN}{BOX['V']}{C.RESET} {C.BOLD}Report status:{C.RESET} Successful")
        print(f"{C.GREEN}{BOX['V']}{C.RESET} {C.BOLD}Report file:{C.RESET} {os.path.basename(saved_path)}")
        print(f"{C.GREEN}{BOX['V']}{C.RESET} {C.BOLD}Report saved to path:{C.RESET} {REPORT_FOLDER}")
        if summary:
            print(f"{C.GREEN}{BOX['V']}{C.RESET}")
            print(f"{C.GREEN}{BOX['V']}{C.RESET} {C.BOLD}Summary{C.RESET}  Total BANs: {C.BOLD}{total}{C.RESET}")
            print(f"{C.GREEN}{BOX['V']}{C.RESET} {C.DIM}Below are the BANs status:{C.RESET}")
            print(f"{C.GREEN}{BOX['V']}{C.RESET}   {C.GREEN}Passed:{C.RESET}   {summary['passed']:>3}  {pct(summary['passed'])}")
            print(f"{C.GREEN}{BOX['V']}{C.RESET}   {C.RED}Failed:{C.RESET}   {summary['failed']:>3}  {pct(summary['failed'])}")
            print(f"{C.GREEN}{BOX['V']}{C.RESET}   {C.YELLOW}Not found:{C.RESET} {summary['not_found']:>3}  {pct(summary['not_found'])}")
            if "tdr_passed" in summary:
                print(f"{C.GREEN}{BOX['V']}{C.RESET} {C.DIM}TDR-wise:{C.RESET} {C.GREEN}Passed:{C.RESET} {summary['tdr_passed']}  {C.RED}Failed:{C.RESET} {summary['tdr_failed']}  {C.YELLOW}Partial:{C.RESET} {summary['tdr_partial']} (Partial = some BANs not in LVT / OOS)")
            if summary.get("per_tdr_count"):
                print(f"{C.GREEN}{BOX['V']}{C.RESET} {C.DIM}One Excel per TDR:{C.RESET} {summary['per_tdr_count']} file(s) saved to {summary.get('per_tdr_folder', '')}")
        print(f"{C.GREEN}{BOX['BL']}{BOX['H'] * (W - 2)}{BOX['BR']}{C.RESET}")
    return all_sources


if __name__ == "__main__":
    main()
