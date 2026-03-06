"""
LVT TDR Delivery Script - Offline CLI for Citrix.

Flow:
  1. Load LVT report -> get all customer IDs from BAN column and LVT status (Pass/Fail).
  2a. Phase 1 – Extract from all input files/sheets: find every 9-digit customer ID and assign TDR (or "No TDR") from section/row (no fixed column/row).
  2b. Save TDR Customer List Excel (Customer ID, TDR Number, Source).
  2c. Phase 2 – Compare LVT IDs with that list -> build merged (Found / Found but no TDR / Not found).
  3. Write mapping Excel (Customer ID, TDR Number, Status, LVT Status, Source File).
  4. Generate INSERT SQL for USCAPP.BAN_MASTER_LIST_LVT and save to report folder (no DB connection).

Config: lvt_tdr_config.json (optional). Prompts for missing values.
After step 3 (mapping Excel), user is asked step-by-step to approve each SQL/DB action.
Dependencies: openpyxl (required); oracledb when running DB/SQL steps.
"""

import json
import os
import queue
import re
import shutil
import sys
import threading
import traceback
from datetime import datetime
from pathlib import Path

# Backend log: always written next to script, so we have a log even if script crashes or runs from another dir
SCRIPT_DIR = Path(__file__).resolve().parent
LOG_DIR = SCRIPT_DIR / "lvt_tdr_logs"

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

CONFIG_NAME = "lvt_tdr_config.json"
SYNTHETIC_PREFIX = "9602"
TDR_PATTERN = re.compile(r"TDR[\s\-_]*(\d{5,6})", re.IGNORECASE)
NINE_DIGIT_PATTERN = re.compile(r"\b(\d{9})\b")
LVT_SHEET_DEFAULT = "BAN Wise Result"
NO_TDR_LABEL = "No TDR"
# Table name for generated INSERT SQL (actual table; no DB connection, we only write .sql file)
BAN_MASTER_TABLE_SQL = "USCAPP.BAN_MASTER_LIST_LVT"


def load_config():
    """Load config from script dir or current dir. Returns dict (possibly empty)."""
    for d in (SCRIPT_DIR, Path.cwd()):
        p = d / CONFIG_NAME
        if p.is_file():
            try:
                with open(p, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                print(f"Warning: Could not load config {p}: {e}")
    return {}


def get_base_folder(config):
    """Resolve base folder: config base_folder, else config downloads_folder, else script_dir/input."""
    base = config.get("base_folder") or config.get("downloads_folder") or ""
    if not base:
        base = SCRIPT_DIR / "input"
    return Path(base).resolve()


def resolve_path(base: Path, path_spec):
    """path_spec can be absolute path string or filename under base."""
    p = Path(path_spec)
    if not p.is_absolute():
        p = base / p
    return p.resolve()


def archive_old_report_files(report_dir):
    """
    Move files in report_dir whose modification date is before today into report_dir/archive.
    Keeps only today's files (by date) in report folder. Creates archive if needed.
    """
    report_dir = Path(report_dir)
    if not report_dir.is_dir():
        return
    today = datetime.now().date()
    archive_dir = report_dir / "archive"
    archive_dir.mkdir(parents=True, exist_ok=True)
    for f in list(report_dir.iterdir()):
        if f.name == "archive" or f.is_dir():
            continue
        try:
            mtime = f.stat().st_mtime
            file_date = datetime.fromtimestamp(mtime).date()
            if file_date < today:
                dest = archive_dir / f.name
                if dest.exists():
                    dest = archive_dir / f"{f.stem}_{int(mtime)}{f.suffix}"
                shutil.move(str(f), str(dest))
        except (OSError, shutil.Error):
            pass


def _list_lvt_files_in_folder(folder):
    """Return list of (full_path, name) for Excel files in folder whose name contains 'LVT' (case-insensitive)."""
    folder = Path(folder)
    if not folder.is_dir():
        return []
    result = []
    for f in folder.iterdir():
        if not f.is_file():
            continue
        if f.suffix.lower() not in (".xlsx", ".xlsm"):
            continue
        if "LVT" in f.name.upper():
            result.append((f.resolve(), f.name))
    return sorted(result, key=lambda x: x[1])


def _list_data_excel_in_folder(folder):
    """Return list of (full_path, name) for Excel files in folder that do NOT contain 'LVT' in name (use as TDR/data input)."""
    folder = Path(folder)
    if not folder.is_dir():
        return []
    result = []
    for f in folder.iterdir():
        if not f.is_file():
            continue
        if f.suffix.lower() not in (".xlsx", ".xlsm"):
            continue
        if "LVT" not in f.name.upper():
            result.append((f.resolve(), f.name))
    return sorted(result, key=lambda x: x[1])


def _list_tdr_files_in_folder(folder):
    """Return list of (full_path, name) for Excel files in folder whose name starts with 'TDR' (case-insensitive)."""
    folder = Path(folder)
    if not folder.is_dir():
        return []
    result = []
    for f in folder.iterdir():
        if not f.is_file():
            continue
        if f.suffix.lower() not in (".xlsx", ".xlsm"):
            continue
        if f.name.upper().startswith("TDR"):
            result.append((f.resolve(), f.name))
    return sorted(result, key=lambda x: x[1])


def _resolve_input_file_list(base_folder, input_file_specs):
    """
    Resolve list of input file paths. If a spec (e.g. 'TDR Data.xlsx') is not found,
    try to find any TDR*.xlsx / LVT*.xlsx in base_folder and use that.
    Returns list of resolved Paths that exist.
    """
    resolved = []
    for spec in input_file_specs:
        path = resolve_path(base_folder, spec)
        if path.is_file():
            resolved.append(path)
            continue
        # Not found: if spec looks like TDR file, look for any TDR*.xlsx in folder
        spec_lower = Path(spec).name.lower()
        if "tdr" in spec_lower or spec_lower.startswith("tdr"):
            tdr_list = _list_tdr_files_in_folder(base_folder)
            if tdr_list:
                resolved.append(tdr_list[0][0])
        else:
            # Could add LVT fallback similarly if needed
            pass
    return resolved


def resolve_lvt_path(base: Path, config_lvt_path, log_fn=None, input_callback=None):
    """
    Resolve LVT report path. If the configured/default path does not exist,
    look for any Excel file starting with 'LVT' in base folder and let user pick or enter path.
    input_callback(prompt, default): if set, used instead of input() for prompts.
    Returns Path to use for LVT report.
    """
    def _input(prompt, default=""):
        if input_callback:
            return (input_callback(prompt, default) or "").strip()
        return input(prompt).strip()
    default_name = config_lvt_path or "LVT_RUN_3Mar_Report.xlsx"
    candidate = resolve_path(base, default_name)
    if candidate.is_file():
        return candidate
    # Not found: look for any Excel with 'LVT' in name in base folder
    lvt_files = _list_lvt_files_in_folder(base)
    if log_fn:
        log_fn(f"[DEBUG] Default LVT path not found. Scanning {base} for Excel files with 'LVT' in name ...")
    if len(lvt_files) == 1:
        if log_fn:
            log_fn(f"[DEBUG] Using single LVT file found: {lvt_files[0][1]}")
        return lvt_files[0][0]
    if len(lvt_files) > 1:
        if log_fn:
            log_fn("LVT report not found at default path. Files with 'LVT' in name in folder:")
            for i, (_, name) in enumerate(lvt_files, 1):
                log_fn(f"  {i}. {name}")
        choice = _input("Enter number or filename to use (or full path): ")
        if not choice:
            raise FileNotFoundError(f"LVT report not found: {candidate}")
        if choice.isdigit():
            idx = int(choice)
            if 1 <= idx <= len(lvt_files):
                return lvt_files[idx - 1][0]
        for full_path, name in lvt_files:
            if name == choice or name.lower() == choice.lower():
                return full_path
        # Treat as path
        p = Path(choice)
        if p.is_absolute() and p.is_file():
            return p
        p = base / choice
        if p.is_file():
            return p.resolve()
    # No LVT files or user entered path
    if log_fn:
        log_fn("Enter path or filename for LVT report (in current folder or full path):")
    choice = _input("LVT report path: ")
    if not choice:
        raise FileNotFoundError(f"LVT report not found: {candidate}")
    p = Path(choice)
    if not p.is_absolute():
        p = base / p
    p = p.resolve()
    if not p.is_file():
        raise FileNotFoundError(f"LVT report not found: {p}")
    return p


# ---------------------------------------------------------------------------
# Step 1: LVT report -> list of all customer IDs from BAN column
# ---------------------------------------------------------------------------

def _cell_value_starts_with(value, prefix):
    if value is None:
        return False
    s = str(value).strip()
    return s.startswith(prefix)


def _find_customer_column_in_lvt(ws):
    """Return 1-based column index for BAN/customer ID. Prefer header row."""
    ban_keywords = ("ban", "bans", "customer", "customer id", "account", "cid")
    for r in range(1, min(ws.max_row + 1, 15)):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None)
        if not row:
            continue
        for c, val in enumerate(row):
            if val is None:
                continue
            if str(val).strip().lower() in ban_keywords:
                return c + 1
    return 1


def _find_status_column_in_lvt(ws):
    """Return 1-based column index for Status (Pass/Fail etc.). Prefer header row."""
    status_keywords = ("status", "result", "lvt", "pass", "fail", "verified", "outcome")
    for r in range(1, min(ws.max_row + 1, 15)):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None)
        if not row:
            continue
        for c, val in enumerate(row):
            if val is None:
                continue
            if str(val).strip().lower() in status_keywords:
                return c + 1
    return 2


def get_synthetic_customer_ids(lvt_path, sheet_name=None, base_folder=None):
    """
    Read LVT Excel and return (list of all customer IDs from BAN column, dict customer_id -> LVT status).
    No filter by prefix; all non-empty BAN/customer values are included. Skips header-like cells.
    sheet_name: default BAN Wise Result. base_folder used if lvt_path is relative.
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl required. pip install openpyxl")
        sys.exit(1)

    path = Path(lvt_path)
    if base_folder and not path.is_absolute():
        path = Path(base_folder) / path
    path = path.resolve()
    if not path.is_file():
        raise FileNotFoundError(f"LVT report not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = sheet_name or LVT_SHEET_DEFAULT
    if sheet_name not in wb.sheetnames:
        ws = wb.active
        sheet_name = ws.title
    else:
        ws = wb[sheet_name]

    ban_col = _find_customer_column_in_lvt(ws)
    status_col = _find_status_column_in_lvt(ws)
    # Skip header row and any header-like cell (so Total BAN = data rows only)
    ban_header_keywords = (
        "ban", "bans", "customer", "customer id", "account", "cid",
        "lgc_customer_id", "customer_id", "status", "lgc customer id",
    )
    customer_ids = []
    seen = set()
    lvt_status = {}
    # Start from row 2 so row 1 is always treated as header (common for LVT sheets)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or ban_col > len(row):
            continue
        val = row[ban_col - 1]
        if val is None:
            continue
        cid = str(val).strip()
        if not cid or cid.lower() in ban_header_keywords:
            continue
        # Skip if value looks like a header (e.g. "customer" + "id", or only letters)
        if "customer" in cid.lower() and "id" in cid.lower():
            continue
        if cid.replace("_", "").replace("-", "").replace(" ", "").isalpha():
            continue
        status_val = row[status_col - 1] if status_col <= len(row) else None
        status_str = str(status_val).strip() if status_val is not None else ""
        # Only count rows with Passed or Failed so Total BAN = Passed + Failed (no header/blank/other)
        if status_str.lower() not in ("passed", "failed"):
            continue
        lvt_status[cid] = status_str
        if cid not in seen:
            seen.add(cid)
            customer_ids.append(cid)
    wb.close()
    return customer_ids, lvt_status


# ---------------------------------------------------------------------------
# Step 2: Search customer IDs in input files; for TDR file resolve TDR section
# ---------------------------------------------------------------------------

def _extract_tdr_from_cell(value):
    if value is None:
        return None
    m = TDR_PATTERN.search(str(value).strip())
    return f"TDR-{m.group(1)}" if m else None


def _get_tdr_section_ranges(ws):
    """Return list of (tdr_id, start_row, end_row_exclusive)."""
    tdr_rows = []
    for row in ws.iter_rows():
        row_idx = next((c.row for c in row if hasattr(c, "row")), None)
        if not row_idx:
            continue
        for cell in row:
            tdr_id = _extract_tdr_from_cell(cell.value)
            if tdr_id:
                tdr_rows.append((row_idx, tdr_id))
                break
    if not tdr_rows:
        return []
    result = []
    for i, (start_row, tdr_id) in enumerate(tdr_rows):
        end_row = tdr_rows[i + 1][0] if i + 1 < len(tdr_rows) else (ws.max_row + 1)
        result.append((tdr_id, start_row, end_row))
    return result


def _row_in_section(row_index, section_ranges):
    """Return tdr_id if row_index is inside a section, else None."""
    for tdr_id, start_row, end_row in section_ranges:
        if start_row <= row_index < end_row:
            return tdr_id
    return None


def _extract_nine_digit_ids_from_cell(value):
    """Return set of 9-digit strings found in cell (number or text). Handles Excel numeric and scientific notation."""
    if value is None:
        return set()
    out = set()
    # Excel often stores IDs as number: int 100084104 or float 100084104.0 / 9.601e+8
    if isinstance(value, (int, float)):
        try:
            n = int(round(value))
            if 100_000_000 <= n <= 999_999_999:
                out.add(str(n))
            return out
        except (ValueError, OverflowError):
            pass
    s = str(value).strip()
    out.update(NINE_DIGIT_PATTERN.findall(s))
    return out


def extract_customer_tdr_from_sheet(ws, sheet_name, file_name, section_ranges):
    """
    Scan entire sheet: find every 9-digit customer ID in any cell.
    For each ID, assign TDR if that row is in a TDR section, else NO_TDR_LABEL.
    Returns list of (customer_id, tdr_value, source_label).
    """
    result = []
    for row in ws.iter_rows():
        row_idx = next((c.row for c in row if hasattr(c, "row")), None)
        if not row_idx:
            continue
        for cell in row:
            for cid in _extract_nine_digit_ids_from_cell(cell.value):
                tdr_id = _row_in_section(row_idx, section_ranges) if section_ranges else None
                tdr_value = tdr_id if tdr_id else NO_TDR_LABEL
                source_label = f"{file_name} | {sheet_name}" if sheet_name else file_name
                result.append((cid, tdr_value, source_label))
    return result


def _extract_from_workbook(wb, path_name, is_tdr):
    """Collect all (cid, tdr_value, source_label) from every sheet of an open workbook."""
    file_rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section_ranges = _get_tdr_section_ranges(ws) if is_tdr else []
        rows = extract_customer_tdr_from_sheet(ws, sheet_name, path_name, section_ranges)
        file_rows.extend(rows)
    return file_rows


def extract_all_customer_tdr_from_files(resolved_paths, base_folder, treat_as_tdr=None, log_fn=None):
    """
    Phase 1: From all sheets of all input files, extract every 9-digit customer ID
    and its TDR (or No TDR). Returns (merged_dict, all_rows_list).
    merged_dict: customer_id -> {"tdr": "TDR-xxx" or "No TDR", "source": "file | sheet"} (one per ID for Mapping).
    all_rows_list: list of (cid, tdr_value, source_label) for every occurrence (so TDR Customer List shows every file).
    If a file yields 0 IDs with read_only, retries with full load for compatibility (e.g. Rate Plan Data_QE.xlsx).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl required. pip install openpyxl")
        sys.exit(1)

    treat_as_tdr = treat_as_tdr or set()
    merged = {}
    all_rows = []  # every (cid, tdr, source) so Rate Plan Data_QE and all files appear in TDR Customer List
    for path in resolved_paths:
        path = Path(path) if not isinstance(path, Path) else path
        if not path.is_file():
            if log_fn:
                log_fn(f"  [SKIP] {path.name}: not a file")
            continue
        try:
            is_tdr = path.name in treat_as_tdr or "tdr" in path.name.lower()
            wb = load_workbook(path, read_only=True, data_only=True)
            file_rows = _extract_from_workbook(wb, path.name, is_tdr)
            wb.close()
            used_fallback = False
            if len(file_rows) == 0:
                wb = load_workbook(path, data_only=True)
                file_rows = _extract_from_workbook(wb, path.name, is_tdr)
                wb.close()
                used_fallback = len(file_rows) > 0
            for cid, tdr_value, source_label in file_rows:
                all_rows.append((cid, tdr_value, source_label))
                prev = merged.get(cid)
                if prev is None:
                    merged[cid] = {"tdr": tdr_value, "source": source_label}
                elif tdr_value != NO_TDR_LABEL and prev["tdr"] == NO_TDR_LABEL:
                    merged[cid] = {"tdr": tdr_value, "source": source_label}
                elif prev["tdr"] == NO_TDR_LABEL and tdr_value != NO_TDR_LABEL:
                    merged[cid] = {"tdr": tdr_value, "source": source_label}
            if log_fn:
                suffix = " (full read)" if used_fallback else ""
                log_fn(f"  {path.name}: {len(file_rows)} customer IDs{suffix}")
        except Exception as e:
            import traceback
            msg = f"[WARN] Skipped file {path.name}: {e}"
            (log_fn or print)(msg)
            if not log_fn:
                traceback.print_exc()
    return merged, all_rows


def write_tdr_customer_list_excel(tdr_list_dict, output_path):
    """Write Phase 1 list: Customer ID, TDR Number, Excel File, Sheet Name."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "TDR Customer List"
    headers = ["Customer ID", "TDR Number", "Excel File", "Sheet Name"]
    ws.append(headers)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c, value=headers[c - 1])
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for i, (cid, info) in enumerate(sorted(tdr_list_dict.items()), start=2):
        source = info.get("source", "")
        if " | " in source:
            excel_file, sheet_name = source.split(" | ", 1)
        else:
            excel_file, sheet_name = source, ""
        row = [cid, info.get("tdr", NO_TDR_LABEL), excel_file, sheet_name]
        ws.append(row)
        for c in range(1, len(row) + 1):
            ws.cell(row=i, column=c).border = thin
    wb.save(output_path)
    return output_path


# Extract 9-digit ID for matching; LVT may have "960212345 Y" or "100084104", TDR may have 960212345
def _core_customer_id(value):
    """Extract 9-digit numeric part from string or number (any customer ID format)."""
    if value is None:
        return None
    s = str(value).strip()
    m = NINE_DIGIT_PATTERN.search(s)
    if m:
        return m.group(1)
    if s.isdigit() and len(s) == 9:
        return s
    return None


def _cell_contains_customer_id(value, customer_id):
    if value is None:
        return False
    s = str(value).strip()
    # Exact / substring match (existing behaviour)
    if s == customer_id or s.startswith(customer_id) or customer_id in s:
        return True
    # Normalized match: compare 9-digit core IDs so "960212345 Y" or "100084104" matches cell value
    core_cid = _core_customer_id(customer_id)
    core_cell = _core_customer_id(value)
    if core_cid and core_cell and core_cid == core_cell:
        return True
    if core_cid and core_cid in s:
        return True
    if core_cell and core_cell in customer_id:
        return True
    # Cell may have multiple IDs (e.g. "960240704 (CC4)\n960240383" in column R); check each token
    for part in re.split(r"[\n,;\t]+", s):
        part = part.strip()
        if not part:
            continue
        if core_cid and core_cid == _core_customer_id(part):
            return True
        if core_cid and core_cid in part:
            return True
    return False


def search_customers_in_sheet(ws, customer_ids, section_ranges, is_tdr_file, sheet_name=""):
    """
    For each customer_id, find if it appears in any cell; if so return (tdr_id or None, status, source_sheet).
    status: "Found" | "Found but no TDR" | "Not found".
    """
    results = {}  # customer_id -> (tdr_id or None, status, source_sheet)
    for cid in customer_ids:
        results[cid] = (None, "Not found", sheet_name)

    for row in ws.iter_rows():
        row_idx = next((c.row for c in row if hasattr(c, "row")), None)
        if not row_idx:
            continue
        for cell in row:
            for cid in customer_ids:
                if results[cid][1] == "Found":
                    continue
                if not _cell_contains_customer_id(cell.value, cid):
                    continue
                tdr_id = _row_in_section(row_idx, section_ranges) if is_tdr_file else None
                if is_tdr_file:
                    status = "Found" if tdr_id else "Found but no TDR"
                    prev_tdr, prev_status, _ = results[cid]
                    if prev_status != "Found" or (tdr_id and not prev_tdr):
                        results[cid] = (tdr_id, status, sheet_name)
                else:
                    results[cid] = (None, "Found", sheet_name)
                break
    return results


def search_all_input_files(input_file_paths, customer_ids, base_folder, treat_as_tdr=None):
    """
    input_file_paths: list of paths (relative to base_folder or absolute).
    treat_as_tdr: set of basenames (e.g. {"TDR Data.xlsx"}) to apply TDR section logic; others just Found/Not found.
    Returns: dict customer_id -> {"tdr_id": str|None, "status": str, "source_file": str} where source_file is "filename | sheetname".
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl required. pip install openpyxl")
        sys.exit(1)

    treat_as_tdr = treat_as_tdr or set()
    merged = {}
    for cid in customer_ids:
        merged[cid] = {"tdr_id": None, "status": "Not found", "source_file": ""}

    for file_spec in input_file_paths:
        path = resolve_path(base_folder, file_spec)
        if not path.is_file():
            print(f"  Skip (not found): {path}")
            continue
        is_tdr = path.name in treat_as_tdr or "tdr" in path.name.lower()
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            section_ranges = _get_tdr_section_ranges(ws) if is_tdr else []
            sheet_results = search_customers_in_sheet(ws, customer_ids, section_ranges, is_tdr, sheet_name=sheet_name)
            for cid, (tdr_id, status, src_sheet) in sheet_results.items():
                if status == "Not found":
                    continue
                source_label = f"{path.name} | {src_sheet}" if src_sheet else path.name
                prev = merged[cid]
                if prev["status"] == "Not found":
                    merged[cid] = {"tdr_id": tdr_id, "status": status, "source_file": source_label}
                elif tdr_id and not prev["tdr_id"]:
                    merged[cid] = {"tdr_id": tdr_id, "status": "Found", "source_file": source_label}
        wb.close()

    return merged


# ---------------------------------------------------------------------------
# Step 3: Write mapping Excel
# ---------------------------------------------------------------------------

# Label for mapping status "Not found" (customer not in any data file) in TDR Summary
NOT_MAPPING_LABEL = "Not mapping with any files"


def _build_tdr_summary_rows_from_mapping(merged, lvt_status):
    """
    Build TDR Summary from mapping (LVT customers only). One row per TDR / No TDR / Not mapping.
    - TDR rows: customers with status Found and a TDR Number.
    - No TDR: customers with status Found but no TDR.
    - Not mapping with any files: customers with status Not found (not in any data file).
    Returns list of (label, total, passed, failed, not_found, status).
    """
    from collections import defaultdict
    lvt_status = lvt_status or {}
    # Group LVT customers (merged) by mapping bucket
    bucket_to_cids = defaultdict(list)
    for cid, info in merged.items():
        status = (info.get("status") or "").strip()
        tdr_id = info.get("tdr_id")
        if status == "Found" and tdr_id:
            bucket_to_cids[tdr_id].append(cid)
        elif status == "Found but no TDR":
            bucket_to_cids[NO_TDR_LABEL].append(cid)
        else:
            bucket_to_cids[NOT_MAPPING_LABEL].append(cid)

    rows = []
    # Sort: TDR numbers first, then "No TDR", then "Not mapping with any files"
    def sort_key(item):
        label = item[0]
        if label == NO_TDR_LABEL:
            return (1, "")
        if label == NOT_MAPPING_LABEL:
            return (2, "")
        return (0, str(label))

    for bucket in sorted(bucket_to_cids.keys(), key=lambda b: sort_key((b,))):
        cids = bucket_to_cids[bucket]
        total = len(cids)
        passed = failed = not_found = 0
        for cid in cids:
            st = (lvt_status.get(cid) or "").strip().lower()
            if st == "passed":
                passed += 1
            elif st == "failed":
                failed += 1
            else:
                not_found += 1
        if failed > 0:
            status = "Failed"
        elif not_found > 0:
            status = "Partial"
        else:
            status = "Passed"
        rows.append((bucket, total, passed, failed, not_found, status))
    return rows


def _build_tdr_summary_rows(tdr_list_dict, merged, lvt_status):
    """
    Build TDR Summary: for each TDR (from tdr_list_dict, exclude No TDR), compute
    Total BANs, Passed, Failed, Not found, TDR Status. Returns list of (tdr, total, passed, failed, not_found, status).
    (Legacy: used when summary was per TDR data; prefer _build_tdr_summary_rows_from_mapping for LVT report.)
    """
    from collections import defaultdict
    lvt_status = lvt_status or {}
    core_to_merged_key = {}
    for mk in merged:
        core = _core_customer_id(mk) or mk
        core_to_merged_key.setdefault(core, mk)
    tdr_to_bans = defaultdict(list)
    for cid, info in tdr_list_dict.items():
        tdr_val = info.get("tdr", NO_TDR_LABEL)
        if tdr_val == NO_TDR_LABEL:
            continue
        tdr_to_bans[tdr_val].append(cid)
    rows = []
    for tdr in sorted(tdr_to_bans.keys()):
        bans = tdr_to_bans[tdr]
        total = len(bans)
        passed = failed = not_found = 0
        for cid in bans:
            merged_key = core_to_merged_key.get(cid)
            if merged_key is None:
                not_found += 1
                continue
            st = (lvt_status.get(merged_key) or "").strip().lower()
            if st == "passed":
                passed += 1
            elif st == "failed":
                failed += 1
            else:
                not_found += 1
        if failed > 0:
            status = "Failed"
        elif not_found > 0:
            status = "Partial"
        else:
            status = "Passed"
        rows.append((tdr, total, passed, failed, not_found, status))
    return rows


def _set_column_widths(ws, widths):
    """Set column widths for a worksheet. widths: list of widths per column (1-based index)."""
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _auto_column_widths(ws, min_width=8, max_width=60):
    """Auto-adjust column widths from cell content (header + data)."""
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    # Approximate: 1 char ~ 1 unit for default font; add 1 for padding
                    length = min(len(str(cell.value)) + 1, max_width)
                    max_len = max(max_len, length)
        ws.column_dimensions[letter].width = min(max(max_len, min_width), max_width)


def write_mapping_excel(merged, output_path, lvt_status=None, tdr_list_dict=None, tdr_list_all_rows=None, include_tdr_customer_list=False):
    """Write one workbook with Mapping and TDR Summary. Optionally add TDR Customer List sheet if include_tdr_customer_list and tdr_list_all_rows given."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    lvt_status = lvt_status or {}
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    wb = Workbook()
    # Sheet order: Mapping (0) first, then TDR Summary (1), then TDR Customer List (2)
    ws = wb.active
    ws.title = "Mapping"
    headers = ["Customer ID", "TDR Number", "Status", "LVT Status", "Excel File", "Sheet Name"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c, value=headers[c - 1])
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    # Sort: first Found, then Found but no TDR, then Not found; within each by TDR number ascending (not found at end)
    _status_order = {"Found": 0, "Found but no TDR": 1, "Not found": 2}
    mapping_sorted = sorted(
        merged.items(),
        key=lambda item: (
            _status_order.get(item[1].get("status") or "Not found", 2),
            item[1].get("tdr_id") or "ZZZ",
            item[0],
        ),
    )
    for i, (cid, info) in enumerate(mapping_sorted, start=2):
        lvt_st = lvt_status.get(cid, "")
        source = info.get("source_file") or ""
        if " | " in source:
            excel_file, sheet_name = source.split(" | ", 1)
        else:
            excel_file, sheet_name = source, ""
        row = [
            cid,
            info.get("tdr_id") or "",
            info.get("status") or "Not found",
            lvt_st,
            excel_file,
            sheet_name,
        ]
        ws.append(row)
        for c in range(1, len(row) + 1):
            ws.cell(row=i, column=c).border = thin
    _auto_column_widths(ws)

    ws_summary = wb.create_sheet(title="TDR Summary", index=1)
    if merged:
        summary_rows = _build_tdr_summary_rows_from_mapping(merged, lvt_status)
        sum_headers = ["TDR", "Total BANs", "Passed", "Failed", "Not found", "TDR Status"]
        ws_summary.append(sum_headers)
        for c in range(1, len(sum_headers) + 1):
            cell = ws_summary.cell(row=1, column=c, value=sum_headers[c - 1])
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
        passed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for i, (tdr, total, passed, failed, not_found, status) in enumerate(summary_rows, start=2):
            ws_summary.append([tdr, total, passed, failed, not_found, status])
            for col in range(1, 7):
                cell = ws_summary.cell(row=i, column=col)
                cell.border = thin
                if col == 6:
                    if status == "Passed":
                        cell.fill = passed_fill
                    elif status == "Failed":
                        cell.fill = failed_fill
                    else:
                        cell.fill = partial_fill
        _auto_column_widths(ws_summary)
    else:
        ws_summary.append(["No TDR summary (no data)"])
        _auto_column_widths(ws_summary)

    if include_tdr_customer_list and tdr_list_all_rows is not None:
        # One row per (customer ID, file, sheet) so every file (e.g. Rate Plan Data_QE.xlsx) appears
        ws_tdr_list = wb.create_sheet(title="TDR Customer List", index=2)
        list_headers = ["Customer ID", "TDR Number", "Excel File", "Sheet Name"]
        ws_tdr_list.append(list_headers)
        for c in range(1, len(list_headers) + 1):
            cell = ws_tdr_list.cell(row=1, column=c, value=list_headers[c - 1])
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
        # Sort by Excel File, Sheet Name, then Customer ID so rows from same file are together
        def _row_key(item):
            cid, tdr_val, src = item
            if " | " in src:
                excel_file, sheet_name = src.split(" | ", 1)
            else:
                excel_file, sheet_name = src, ""
            return (excel_file, sheet_name, cid)
        for i, (cid, tdr_val, source_label) in enumerate(sorted(tdr_list_all_rows, key=_row_key), start=2):
            if " | " in source_label:
                excel_file, sheet_name = source_label.split(" | ", 1)
            else:
                excel_file, sheet_name = source_label, ""
            row = [cid, tdr_val, excel_file, sheet_name]
            ws_tdr_list.append(row)
            for c in range(1, len(row) + 1):
                ws_tdr_list.cell(row=i, column=c).border = thin
        _auto_column_widths(ws_tdr_list)

    wb.save(output_path)
    return output_path


def write_tdr_list_only_excel(tdr_list_all_rows, output_path):
    """Write a single-sheet Excel with TDR Customer List (Customer ID, TDR Number, Excel File, Sheet Name)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "TDR Customer List"
    list_headers = ["Customer ID", "TDR Number", "Excel File", "Sheet Name"]
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(list_headers)
    for c in range(1, len(list_headers) + 1):
        cell = ws.cell(row=1, column=c, value=list_headers[c - 1])
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    def _row_key(item):
        cid, tdr_val, src = item
        if " | " in src:
            excel_file, sheet_name = src.split(" | ", 1)
        else:
            excel_file, sheet_name = src, ""
        return (excel_file, sheet_name, cid)
    for i, (cid, tdr_val, source_label) in enumerate(sorted(tdr_list_all_rows, key=_row_key), start=2):
        if " | " in source_label:
            excel_file, sheet_name = source_label.split(" | ", 1)
        else:
            excel_file, sheet_name = source_label, ""
        row = [cid, tdr_val, excel_file, sheet_name]
        ws.append(row)
        for c in range(1, len(row) + 1):
            ws.cell(row=i, column=c).border = thin
    _auto_column_widths(ws)
    wb.save(output_path)
    return output_path


def run_tdr_list_only(data_paths, output_path, log_fn=None):
    """
    TDR data analysis only: extract customer IDs + TDR from data files and write a single-sheet
    Excel (TDR Customer List). No LVT, no mapping, no INSERT SQL.
    Returns path to the written Excel, or None if no data.
    """
    log = log_fn or (lambda msg: None)
    data_paths = [Path(p) for p in data_paths if Path(p).is_file()]
    if not data_paths:
        log("No data files.")
        return None
    base = data_paths[0].parent
    treat_tdr = {Path(p).name for p in data_paths}
    tdr_list_dict, tdr_list_all_rows = extract_all_customer_tdr_from_files(
        data_paths, base, treat_as_tdr=treat_tdr, log_fn=log
    )
    if not tdr_list_all_rows:
        log("No customer IDs extracted from the data files.")
        return None
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_tdr_list_only_excel(tdr_list_all_rows, output_path)
    log(f"TDR Customer List: {len(tdr_list_all_rows)} rows written.")
    return output_path


# ---------------------------------------------------------------------------
# Step 4: Generate INSERT SQL file (no DB); table name in SQL = BAN_MASTER_TABLE_SQL
# ---------------------------------------------------------------------------

def get_db_connection(config):
    """Return oracledb connection from config or env."""
    try:
        import oracledb
    except ImportError:
        print("Error: oracledb required. pip install oracledb")
        sys.exit(1)

    # Prefer config; fallback env
    dsn = config.get("db_dsn") or os.environ.get("ORACLE_DSN")
    user = config.get("db_user") or os.environ.get("ORACLE_USER")
    password = config.get("db_password") or os.environ.get("ORACLE_PASSWORD")
    if not all((dsn, user, password)):
        print("DB connection: set db_dsn, db_user, db_password in config or ORACLE_DSN, ORACLE_USER, ORACLE_PASSWORD env.")
        return None
    return oracledb.connect(user=user, password=password, dsn=dsn)


def _sql_escape(s):
    """Escape single quotes for Oracle SQL literal ('' for ')."""
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def write_insert_sql_file(rows_to_insert, owner, requestor, default_tdr_id, output_path, table_name=None, log=None):
    """
    Generate INSERT statements for USCAPP.BAN_MASTER_LIST_LVT and write to a .sql file.
    No DB connection. table_name defaults to BAN_MASTER_TABLE_SQL.
    Returns path to the written file.
    """
    table_name = table_name or BAN_MASTER_TABLE_SQL
    if not rows_to_insert:
        if log:
            log("No rows to write for this INSERT file.")
        return None
    owner_sql = _sql_escape(owner) if owner else "NULL"
    requestor_sql = _sql_escape(requestor) if requestor else "NULL"
    lines = [
        f"-- INSERT statements for {table_name}",
        f"-- Generated {datetime.now().isoformat()}",
        f"-- OWNER={owner or 'NULL'}, REQUESTOR={requestor or 'NULL'}",
        "",
    ]
    count = 0
    for row in rows_to_insert:
        # Support (customer_id, tdr_id) or (customer_id, tdr_id, status_label)
        if len(row) == 3:
            customer_id, tdr_id, status_label = row
        else:
            customer_id, tdr_id = row
            status_label = None
        tdr = tdr_id or default_tdr_id
        if not tdr:
            continue
        cid_sql = _sql_escape(str(customer_id))
        tdr_sql = _sql_escape(tdr)
        status_sql = _sql_escape(status_label) if status_label else "'AVAILABLE'"
        stmt = (
            f"INSERT INTO {table_name}\n"
            " ( CUSTOMER_ID, WAS_USED, TDR_ID, OWNER, DO_NOT_USE, STATUS, DELIVERED_DATE, LOAD_DATE, REQUESTOR, FAILED_RULES, LAST_ACS_RUN_DATE )\n"
            f" VALUES ( {cid_sql}, NULL, {tdr_sql}, {owner_sql}, NULL, {status_sql}, NULL, SYSDATE, {requestor_sql}, NULL, NULL );\n"
        )
        lines.append(stmt)
        count += 1
    output_path = Path(output_path)
    output_path.write_text("\n".join(lines), encoding="utf-8")
    if log:
        log(f"Wrote {count} INSERT statements to {output_path}")
    return output_path


def run_lvt_tdr_from_paths(
    lvt_path,
    data_paths,
    output_dir,
    lvt_sheet_name=None,
    owner=None,
    requestor=None,
    default_tdr_id=None,
    log_fn=None,
    log_paths=True,
):
    """
    Run LVT TDR pipeline from given file paths (for Streamlit or other callers).
    No config, no prompts. Returns (report_excel_path, synth_insert_sql_path, prod_insert_sql_path, summary_dict).
    - lvt_path: Path to LVT Excel. data_paths: list of Paths to data Excel files.
    - output_dir: where to write report and SQL.
    - log_paths=False: do not log file paths (for UI).
    """
    _raw_log = log_fn or (lambda msg: None)
    if not log_paths:
        import re
        def log(msg):
            if "Report:" in msg and (".xlsx" in msg or "/" in msg or "\\" in msg):
                _raw_log("Report ready for download.")
            elif "Wrote " in msg and " INSERT " in msg and " to " in msg:
                m = re.search(r"Wrote (\d+) INSERT", msg)
                n = m.group(1) if m else "0"
                _raw_log(f"INSERT SQL ready ({n} statements) for download.")
            else:
                _raw_log(msg)
    else:
        log = _raw_log
    lvt_path = Path(lvt_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    lvt_sheet_name = lvt_sheet_name or LVT_SHEET_DEFAULT
    data_paths = [Path(p) for p in data_paths if Path(p).is_file()]
    treat_tdr = {Path(p).name for p in data_paths}

    customer_ids, lvt_status = get_synthetic_customer_ids(lvt_path, sheet_name=lvt_sheet_name, base_folder=lvt_path.parent)
    log(f"Step 1: {len(customer_ids)} customer IDs from LVT.")

    if not data_paths:
        tdr_list_dict = {}
        tdr_list_all_rows = []
        log("No data files.")
    else:
        base = data_paths[0].parent
        tdr_list_dict, tdr_list_all_rows = extract_all_customer_tdr_from_files(
            data_paths, base, treat_as_tdr=treat_tdr, log_fn=log
        )
        log(f"Step 2a: {len(tdr_list_dict)} unique IDs from {len(tdr_list_all_rows)} rows.")

    merged = {}
    for cid in customer_ids:
        core = _core_customer_id(cid) or cid
        info = tdr_list_dict.get(core) or tdr_list_dict.get(cid)
        if info is None:
            merged[cid] = {"tdr_id": None, "status": "Not found", "source_file": ""}
        else:
            tdr_val = info.get("tdr", NO_TDR_LABEL)
            tdr_id = None if tdr_val == NO_TDR_LABEL else tdr_val
            status = "Found" if tdr_id else "Found but no TDR"
            merged[cid] = {"tdr_id": tdr_id, "status": status, "source_file": info.get("source", "")}

    report_excel = output_dir / f"LVT_TDR_Report_{ts}.xlsx"
    write_mapping_excel(
        merged, report_excel, lvt_status=lvt_status, tdr_list_dict=tdr_list_dict,
        tdr_list_all_rows=tdr_list_all_rows, include_tdr_customer_list=False,
    )
    log(f"Report: {report_excel}")

    # Build summary for UI (BAN wise + TDR wise)
    total = len(customer_ids)
    passed = sum(1 for c in customer_ids if (lvt_status.get(c) or "").strip().lower() == "passed")
    failed = sum(1 for c in customer_ids if (lvt_status.get(c) or "").strip().lower() == "failed")
    not_found = sum(1 for c in merged.values() if (c.get("status") or "") == "Not found")
    summary_rows = _build_tdr_summary_rows_from_mapping(merged, lvt_status)
    tdr_passed = sum(1 for r in summary_rows if r[5] == "Passed")
    tdr_failed = sum(1 for r in summary_rows if r[5] == "Failed")
    tdr_partial = sum(1 for r in summary_rows if r[5] == "Partial")
    summary = {
        "total": total, "passed": passed, "failed": failed, "not_found": not_found,
        "tdr_passed": tdr_passed, "tdr_failed": tdr_failed, "tdr_partial": tdr_partial,
    }

    # Build INSERT rows: everyone from LVT list; TDR from mapping or default_tdr_id.
    # Bifurcation: customer_id starting with "960" → synthetic SQL; all others → production SQL.
    rows_all = []
    for cid in sorted(customer_ids):
        info = merged.get(cid, {})
        tdr_id = info.get("tdr_id")  # None for Found but no TDR or Not found → use default_tdr_id in write
        raw_lvt = (lvt_status.get(cid) or "").strip()
        lvt_st = raw_lvt.lower()
        status_label = "Passed LVT" if lvt_st == "passed" else "Failed LVT"
        rows_all.append((cid, tdr_id, status_label))

    cid_str = lambda c: str(c).strip()
    rows_synth = [r for r in rows_all if cid_str(r[0]).startswith("960")]
    rows_prod = [r for r in rows_all if not cid_str(r[0]).startswith("960")]

    synth_sql_path = output_dir / f"INSERT_BAN_MASTER_LIST_LVT_SYNTH_{ts}.sql"
    prod_sql_path = output_dir / f"INSERT_BAN_MASTER_LIST_LVT_PRODUCTION_{ts}.sql"
    write_insert_sql_file(
        rows_synth, owner, requestor, default_tdr_id,
        synth_sql_path, table_name=BAN_MASTER_TABLE_SQL, log=log
    )
    write_insert_sql_file(
        rows_prod, owner, requestor, default_tdr_id,
        prod_sql_path, table_name=BAN_MASTER_TABLE_SQL, log=log
    )
    return report_excel, synth_sql_path, prod_sql_path, summary


# ---------------------------------------------------------------------------
# Step 5 & 6: Run SQL files; Pre_load result -> Excel
# ---------------------------------------------------------------------------

def run_sql_file(conn, sql_path, log=None):
    """Execute statements in file (split by ;). Log each error but continue."""
    path = Path(sql_path)
    if not path.is_file():
        if log:
            log(f"SQL file not found: {path}")
        return False
    text = path.read_text(encoding="utf-8", errors="replace")
    statements = [
        s.strip() for s in text.split(";")
        if s.strip() and not s.strip().startswith("--")
    ]
    cursor = conn.cursor()
    ok = True
    for i, stmt in enumerate(statements):
        if not stmt:
            continue
        try:
            cursor.execute(stmt)
            if log:
                log(f"  Statement {i+1} OK")
        except Exception as e:
            if log:
                log(f"  Statement {i+1} failed: {e}")
            ok = False
    conn.commit()
    cursor.close()
    return ok


def run_query_and_save_to_excel(conn, sql_path, output_excel_path, log=None):
    """Run single query from file, fetch all rows, write to Excel."""
    path = Path(sql_path)
    if not path.is_file():
        if log:
            log(f"Query file not found: {path}")
        return False
    text = path.read_text(encoding="utf-8", errors="replace").strip()
    # Use last statement if multiple (e.g. only SELECT)
    parts = [s.strip() for s in text.split(";") if s.strip()]
    sql = parts[-1] if parts else text
    cursor = conn.cursor()
    try:
        cursor.execute(sql)
        rows = cursor.fetchall()
        col_names = [d[0] for d in cursor.description]
    except Exception as e:
        if log:
            log(f"Query failed: {e}")
        cursor.close()
        return False
    cursor.close()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pre_load_Result"
    ws.append(col_names)
    for row in rows:
        ws.append(list(row))
    wb.save(output_excel_path)
    if log:
        log(f"Saved {len(rows)} rows to {output_excel_path}")
    return True


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(session_log_file=None, base_override=None, gui_log_callback=None, gui_ask_yes_no=None, gui_input=None):
    """
    base_override: if set (e.g. from GUI), use this as base folder instead of config.
    gui_log_callback(msg): if set, called for each log line (e.g. to show in GUI).
    gui_ask_yes_no(prompt, default): if set, used instead of console input.
    gui_input(prompt): if set, used instead of input() for text prompts.
    """
    # Backend log: create log dir next to script and write everything here (even if we crash later)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file_owned = False
    if session_log_file is None:
        session_log_path = LOG_DIR / f"lvt_tdr_{ts}.log"
        session_log_file = open(session_log_path, "w", encoding="utf-8")
        log_file_owned = True
        session_log_file.write(f"=== LVT TDR Delivery started at {datetime.now().isoformat()} ===\n")
        session_log_file.write(f"Python: {sys.version}\n")
        session_log_file.write(f"CWD: {Path.cwd()}\n")
        session_log_file.flush()

    log_lines = []

    def log(msg):
        print(msg)
        log_lines.append(msg)
        try:
            session_log_file.write(msg + "\n")
            session_log_file.flush()
        except Exception:
            pass
        if gui_log_callback:
            try:
                gui_log_callback(msg)
            except Exception:
                pass

    def input_fn(prompt, default=""):
        if gui_input:
            return (gui_input(prompt, default) or "").strip() or default
        return (input(prompt).strip() or default)

    log("[DEBUG] Config loading...")
    config = load_config()
    base = Path(base_override).resolve() if base_override else get_base_folder(config)
    log(f"[DEBUG] Base folder (input): {base}")
    output_folder = config.get("output_folder")
    if output_folder:
        out_dir = Path(output_folder)
        if not out_dir.is_absolute():
            out_dir = base / out_dir
    else:
        # When base is "input" folder, put report next to it (sibling) so output is not inside input
        out_dir = (base.parent / "report") if base.name.lower() == "input" else (base / "report")
    out_dir.mkdir(parents=True, exist_ok=True)
    archive_old_report_files(out_dir)
    log_path = LOG_DIR / f"lvt_tdr_run_{ts}.log"
    log(f"Log file (this run): {log_path}  (run log in logs folder)")
    log(f"Backend log dir: {LOG_DIR}")

    # --- Step 1: LVT synthetic customers ---
    log("[DEBUG] Step 1: Resolving LVT path...")
    lvt_path = config.get("lvt_report_path") or "LVT_RUN_3Mar_Report.xlsx"
    try:
        lvt_full = resolve_lvt_path(base, lvt_path, log_fn=log, input_callback=input_fn)
    except FileNotFoundError as e:
        log(f"[ERROR] {e}")
        raise
    log(f"[DEBUG] LVT full path: {lvt_full} (exists: {lvt_full.is_file()})")
    lvt_sheet = config.get("lvt_sheet_name") or LVT_SHEET_DEFAULT
    try:
        customer_ids, lvt_status = get_synthetic_customer_ids(lvt_full, sheet_name=lvt_sheet, base_folder=base)
    except FileNotFoundError as e:
        log(f"[ERROR] {e}")
        raise
    except Exception as e:
        log(f"[ERROR] Step 1 failed: {e}")
        traceback.print_exc()
        raise
    log(f"Step 1: Found {len(customer_ids)} customer IDs from LVT sheet in {lvt_full.name}")

    # --- Step 2a (Phase 1): Extract ALL customer IDs + TDR from every sheet (no fixed column/row) ---
    # Default: all Excel in input folder except files with 'LVT' in name = TDR/data input.
    log("[DEBUG] Step 2a: Extracting all 9-digit customer IDs and TDR from input files (all sheets)...")
    config_input_files = config.get("input_files")
    if config_input_files and len(config_input_files) > 0:
        resolved_input = _resolve_input_file_list(base, config_input_files)
        treat_tdr = set(config.get("tdr_file_names") or []) or {p.name for p in resolved_input}
    else:
        data_files = _list_data_excel_in_folder(base)
        resolved_input = [p for (p, _) in data_files]
        treat_tdr = {n for (_, n) in data_files}
        log(f"[DEBUG] Using all Excel in input folder (except LVT) as data: {[n for (_, n) in data_files]}")
    if not resolved_input:
        log("[WARN] No data input files found. Put Excel files (non-LVT) in the input folder.")
        tdr_list_dict = {}
        tdr_list_all_rows = []
    else:
        log(f"[DEBUG] Data input files ({len(resolved_input)}): {[p.name for p in resolved_input]}")
        try:
            tdr_list_dict, tdr_list_all_rows = extract_all_customer_tdr_from_files(
                resolved_input, base, treat_as_tdr=treat_tdr, log_fn=log
            )
        except Exception as e:
            log(f"[ERROR] Step 2a failed: {e}")
            traceback.print_exc()
            raise
        log(f"Step 2a: Extracted {len(tdr_list_dict)} unique customer IDs from {len(tdr_list_all_rows)} rows in {len(resolved_input)} file(s).")

    # --- Step 2b (Phase 2): Compare LVT customer IDs with TDR list → build merged ---
    log("[DEBUG] Step 2b: Comparing LVT customer IDs with TDR list...")
    merged = {}
    for cid in customer_ids:
        core = _core_customer_id(cid) or cid
        info = tdr_list_dict.get(core) or tdr_list_dict.get(cid)
        if info is None:
            merged[cid] = {"tdr_id": None, "status": "Not found", "source_file": ""}
        else:
            tdr_val = info.get("tdr", NO_TDR_LABEL)
            tdr_id = None if tdr_val == NO_TDR_LABEL else tdr_val
            status = "Found" if tdr_id else "Found but no TDR"
            merged[cid] = {"tdr_id": tdr_id, "status": status, "source_file": info.get("source", "")}
    log(f"Step 2b: Compared {len(customer_ids)} LVT IDs with TDR list.")

    # --- Step 3: Write single report Excel (TDR Summary + Mapping + TDR Customer List sheets) ---
    log("[DEBUG] Step 3: Writing report Excel...")
    report_excel = out_dir / f"LVT_TDR_Report_{ts}.xlsx"
    try:
        write_mapping_excel(
            merged, report_excel, lvt_status=lvt_status, tdr_list_dict=tdr_list_dict,
            tdr_list_all_rows=tdr_list_all_rows,
        )
    except Exception as e:
        log(f"[ERROR] Step 3 failed: {e}")
        traceback.print_exc()
        raise
    log(f"Step 3: Report saved to {report_excel} (sheets: TDR Summary, Mapping, TDR Customer List)")

    # Summary: Found / Found but no TDR / Not found (so user sees why insert may be empty)
    count_found = sum(1 for c in merged.values() if c.get("status") == "Found" and c.get("tdr_id"))
    count_found_no_tdr = sum(1 for c in merged.values() if c.get("status") == "Found but no TDR")
    count_not_found = sum(1 for c in merged.values() if c.get("status") == "Not found")
    log(f"Mapping summary:  Found (with TDR): {count_found}  |  Found but no TDR: {count_found_no_tdr}  |  Not found in TDR: {count_not_found}")

    # Build list of rows to insert: only "Found" or "Found but no TDR", and only LVT status = Passed
    rows_to_insert = []
    for cid in sorted(merged.keys()):
        info = merged[cid]
        status = info.get("status") or "Not found"
        tdr_id = info.get("tdr_id")
        lvt_st = (lvt_status.get(cid) or "").strip().lower()
        if lvt_st != "passed":
            continue
        if status == "Found" and tdr_id:
            rows_to_insert.append((cid, tdr_id))
        elif status == "Found but no TDR":
            rows_to_insert.append((cid, None))

    def _cli_ask_yes_no(prompt, default=False):
        suffix = " [Y/n]: " if default else " [y/N]: "
        while True:
            r = input(prompt + suffix).strip().lower()
            if not r:
                return default
            if r in ("y", "yes"):
                return True
            if r in ("n", "no"):
                return False

    def ask_yes_no(prompt, default=False):
        if gui_ask_yes_no:
            return gui_ask_yes_no(prompt, default)
        return _cli_ask_yes_no(prompt, default)


    # --- Ask user to check Excel, then proceed step-by-step ---
    log("")
    log("Please check the mapping Excel and validate. INSERT SQL will be saved in the report folder (no DB connection).")
    if not rows_to_insert:
        log("No rows eligible for insert (only LVT Passed + Found or Found but no TDR).")

    conn = None

    # --- Step 4: Generate INSERT SQL file (no DB connection; save to report folder) ---
    log("")
    conn = None
    if rows_to_insert:
        no_tdr_count = sum(1 for _, tdr_id in rows_to_insert if tdr_id is None)
        log(f"INSERT SQL: {len(rows_to_insert)} rows for {BAN_MASTER_TABLE_SQL}:")
        if no_tdr_count:
            log(f"  ({no_tdr_count} have no TDR; default TDR will be used in SQL if provided.)")
        for i, (cid, tdr_id) in enumerate(rows_to_insert[:20], 1):
            tdr = tdr_id or "(use default)"
            log(f"  {i}. Customer ID: {cid}  ->  TDR_ID: {tdr}")
        if len(rows_to_insert) > 20:
            log(f"  ... and {len(rows_to_insert) - 20} more.")
        owner = config.get("owner") or input_fn("Enter OWNER for INSERT SQL (or leave blank): ") or None
        requestor = config.get("requestor") or input_fn("Enter REQUESTOR (or leave blank): ") or None
        default_tdr_id = config.get("default_tdr_id_when_no_mapping")
        has_no_tdr = any(tdr_id is None for _, tdr_id in rows_to_insert)
        if has_no_tdr and not default_tdr_id:
            default_tdr_id = input_fn("Some rows have no TDR. Enter default TDR_ID for SQL (or leave blank to skip those rows): ") or None
        insert_sql_path = out_dir / f"INSERT_BAN_MASTER_LIST_LVT_{ts}.sql"
        try:
            write_insert_sql_file(
                rows_to_insert, owner, requestor, default_tdr_id,
                insert_sql_path, table_name=BAN_MASTER_TABLE_SQL, log=log
            )
        except Exception as e:
            log(f"[WARN] Could not write INSERT SQL file: {e}")
    else:
        log("INSERT SQL: No rows (only LVT Passed + Found or Found but no TDR).")

    # --- Step 5: R2 queries (disabled; no DB connection) ---
    if conn:
        log("")
        r2_sql = config.get("r2_sql_path") or "R2_DATA_DELIVERY_QUERIES_V2.sql"
        r2_full = resolve_path(base, r2_sql)
        if ask_yes_no(f"Ready to run R2 queries from {r2_full.name}?"):
            log("Running R2_DATA_DELIVERY_QUERIES_V2.sql ...")
            run_sql_file(conn, r2_full, log=log)
            log("R2 queries run finished.")
        else:
            log("R2 queries skipped by user.")

    # --- Step 6: Pre_load query -> Excel (with approval) ---
    if conn:
        log("")
        preload_sql = config.get("preload_sql_path") or "Pre_load_updated_query_v3.sql"
        preload_full = resolve_path(base, preload_sql)
        preload_excel = out_dir / f"Pre_load_result_{ts}.xlsx"
        if ask_yes_no("Ready to run Pre_load query and save result to Excel?"):
            log("Running Pre_load_updated_query_v3.sql and saving result ...")
            ok = run_query_and_save_to_excel(conn, preload_full, preload_excel, log=log)
            if ok:
                log(f"Final result saved to: {preload_excel}")
            else:
                log("Pre_load query failed or file not found.")
        else:
            log("Pre_load step skipped by user.")
        conn.close()

    # Write run log to output folder and close backend log if we opened it
    try:
        with open(log_path, "w", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        log(f"[DEBUG] Run log also saved to: {log_path}")
    except Exception as e:
        log(f"[WARN] Could not write run log to {log_path}: {e}")

    log("Done.")
    if log_file_owned and session_log_file is not None:
        try:
            session_log_file.close()
        except Exception:
            pass


def _run_with_logging_and_pause():
    """Run main with backend log and pause on exit so window does not close immediately."""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    session_log_path = LOG_DIR / f"lvt_tdr_{ts}.log"
    try:
        with open(session_log_path, "w", encoding="utf-8") as session_log:
            session_log.write(f"=== LVT TDR Delivery started at {datetime.now().isoformat()} ===\n")
            session_log.write(f"Python: {sys.version}\n")
            session_log.write(f"CWD: {Path.cwd()}\n")
            session_log.write(f"Script: {Path(__file__).resolve()}\n")
            session_log.flush()
            try:
                main(session_log_file=session_log)
            except SystemExit as e:
                session_log.write(f"\n[EXIT] SystemExit: {e}\n")
                session_log.flush()
                raise
            except Exception:
                tb = traceback.format_exc()
                session_log.write("\n\n!!! EXCEPTION !!!\n")
                session_log.write(tb)
                session_log.flush()
                print("\n\n!!! EXCEPTION !!!\n")
                print(tb)
                print(f"\nLog saved to: {session_log_path}")
                input("\nPress Enter to close...")
                raise
    except Exception:
        raise
    # Normal exit: pause so user can read output when double-clicking
    print(f"\nLog file: {session_log_path}")
    input("Press Enter to close...")


def run_gui():
    """Launch Tkinter GUI: select input folder, run, see log in window."""
    try:
        import tkinter as tk
        from tkinter import ttk, filedialog, messagebox, simpledialog
    except ImportError:
        print("Tkinter not available. Run with --cli for console mode.")
        _run_with_logging_and_pause()
        return

    # Ensure default input folder exists next to script
    input_default = SCRIPT_DIR / "input"
    input_default.mkdir(parents=True, exist_ok=True)

    root = tk.Tk()
    root.title("LVT TDR Delivery")
    root.minsize(500, 400)
    root.geometry("700x500")

    # Queues: worker -> main (log lines, prompt requests); main -> worker (prompt responses)
    log_queue = queue.Queue()
    request_queue = queue.Queue()
    response_queue = queue.Queue()

    # Top: input folder
    top = ttk.Frame(root, padding=10)
    top.pack(fill="x")
    ttk.Label(top, text="Input folder (LVT + TDR Excel files):").grid(row=0, column=0, sticky="w", pady=2)
    folder_var = tk.StringVar(value=str(input_default))
    entry = ttk.Entry(top, textvariable=folder_var, width=70)
    entry.grid(row=1, column=0, sticky="ew", padx=(0, 5), pady=2)

    def browse():
        start = folder_var.get() or str(SCRIPT_DIR)
        path = filedialog.askdirectory(initialdir=start, title="Select input folder")
        if path:
            folder_var.set(path)

    ttk.Button(top, text="Browse...", command=browse).grid(row=1, column=1, pady=2)
    top.columnconfigure(0, weight=1)

    # Log area
    log_frame = ttk.LabelFrame(root, text="Log", padding=5)
    log_frame.pack(fill="both", expand=True, padx=10, pady=5)
    log_text = tk.Text(log_frame, wrap="word", height=18, state="disabled", font=("Consolas", 9))
    log_text.pack(fill="both", expand=True)
    scroll = ttk.Scrollbar(log_frame, command=log_text.yview)
    scroll.pack(side="right", fill="y")
    log_text.config(yscrollcommand=scroll.set)

    # Run and Close buttons
    btn_frame = ttk.Frame(root)
    btn_frame.pack(pady=10)
    run_btn = ttk.Button(btn_frame, text="Run", command=lambda: None)
    run_btn.pack(side="left", padx=5)
    close_btn = ttk.Button(btn_frame, text="Close", command=root.destroy)
    close_btn.pack(side="left", padx=5)

    def gui_log(msg):
        log_queue.put(("log", msg))

    def gui_ask_yes_no(prompt, default=False):
        request_queue.put(("ask_yes_no", prompt, default))
        return response_queue.get()

    def gui_input(prompt, default=""):
        request_queue.put(("input", prompt, default))
        return response_queue.get()

    def process_queues():
        # Drain log queue
        while True:
            try:
                item = log_queue.get_nowait()
                if item[0] == "log":
                    log_text.config(state="normal")
                    log_text.insert("end", item[1] + "\n")
                    log_text.see("end")
                    log_text.config(state="disabled")
            except queue.Empty:
                break
        # One prompt request (so dialog appears in order)
        try:
            req = request_queue.get_nowait()
            if req[0] == "ask_yes_no":
                # askyesno expects default="yes" or "no", not boolean
                default_opt = "yes" if req[2] else "no"
                result = messagebox.askyesno("Confirm", req[1], default=default_opt)
                response_queue.put(result)
            elif req[0] == "input":
                result = simpledialog.askstring("Input", req[1], initialvalue=req[2] if len(req) > 2 else "")
                response_queue.put(result if result is not None else "")
        except queue.Empty:
            pass
        root.after(150, process_queues)

    def do_run():
        run_btn.config(state="disabled")
        base_path = folder_var.get().strip()
        if not base_path:
            messagebox.showwarning("Input folder", "Please select an input folder.")
            run_btn.config(state="normal")
            return
        if not Path(base_path).is_dir():
            messagebox.showerror("Input folder", f"Folder does not exist:\n{base_path}")
            run_btn.config(state="normal")
            return

        def worker():
            try:
                main(
                    session_log_file=None,
                    base_override=base_path,
                    gui_log_callback=gui_log,
                    gui_ask_yes_no=gui_ask_yes_no,
                    gui_input=gui_input,
                )
                log_queue.put(("log", "Done."))
            except Exception as e:
                log_queue.put(("log", f"!!! ERROR: {e}"))
                log_queue.put(("log", traceback.format_exc()))
            finally:
                root.after(0, lambda: run_btn.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    run_btn.config(command=do_run)
    root.after(150, process_queues)
    root.mainloop()


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1].strip().lower() in ("--cli", "-c"):
        _run_with_logging_and_pause()
    else:
        run_gui()
